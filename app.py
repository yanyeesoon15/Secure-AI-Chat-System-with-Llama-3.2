from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify, send_file
from werkzeug.security import generate_password_hash, check_password_hash
import sqlite3
import os
import secrets
import requests
from datetime import datetime, timedelta
import threading
import time
import traceback
import json
import re
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.colors import HexColor
from reportlab.lib.units import inch
from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle, Paragraph, 
                                 Spacer, PageBreak)
import PyPDF2
import docx
import csv
from Crypto.Cipher import AES
from Crypto.Util.Padding import pad, unpad
import base64
import hashlib
from functools import wraps

# NEW: Import Vector Database
import chromadb
from chromadb.config import Settings
import uuid

app = Flask(__name__)

# SECURITY CONFIGURATION 
app.secret_key = os.environ.get('SECRET_KEY', secrets.token_hex(32))

app.config.update(
    SESSION_COOKIE_SECURE=False,
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE='Lax',
    PERMANENT_SESSION_LIFETIME=timedelta(hours=8)
)

UPLOAD_FOLDER = "uploads"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

MAX_FILE_SIZE = 5 * 1024 * 1024

# GLOBAL SECURITY HOOK 
@app.before_request
def sync_session_access_level():
    """
    Runs on EVERY request to ensure access is synced with the database
    using strictly UTC time to prevent timezone desynchronization.
    """
    if "user_id" in session:
        # 1. Force the database cleanup right now using UTC
        try:
            with get_db() as conn:
                # Use 'utc' modifier to guarantee SQLite uses UTC for comparison
                conn.execute("UPDATE active_access SET is_active=0 WHERE user_id=? AND expires_at <= datetime('now', 'utc') AND is_active=1", (session["user_id"],))
                conn.execute("UPDATE access_requests SET status='expired' WHERE user_id=? AND status='approved' AND expires_at <= datetime('now', 'utc')", (session["user_id"],))
                conn.commit()
        except Exception:
            pass 

        # 2. Sync the cookie to the new database reality
        # This prevents the user from having "Medium" access in their cookie
        # if the database has already revoked it.
        live_access_level = get_effective_access_level(session["user_id"])
        if session.get("access_level") != live_access_level:
            session["access_level"] = live_access_level

# VECTOR DATABASE SETUP 
CHROMA_DB_DIR = "chroma_db"
os.makedirs(CHROMA_DB_DIR, exist_ok=True)
chroma_client = chromadb.PersistentClient(path=CHROMA_DB_DIR)
# Initialize collection for document chunks
vector_collection = chroma_client.get_or_create_collection(name="secure_corporate_docs")

def chunk_text(text, chunk_size=300, overlap=50):
    """Splits text into smaller, overlapping chunks for accurate vector search."""
    if not text: return []
    words = text.split()
    chunks = []
    for i in range(0, len(words), max(1, chunk_size - overlap)):
        chunk = " ".join(words[i:i + chunk_size])
        if chunk.strip():
            chunks.append(chunk)
    return chunks

def sync_chromadb():
    """Syncs existing SQLite documents to ChromaDB on startup."""
    print("Syncing SQLite documents to Vector Database...")
    try:
        with get_db() as conn:
            docs = conn.execute("SELECT id, filename, classification, content_encrypted, content, filepath FROM documents").fetchall()
        
        # Get existing IDs in Chroma to avoid duplicates
        existing_data = vector_collection.get()
        existing_doc_ids = set()
        if existing_data and existing_data['metadatas']:
            existing_doc_ids = set([meta['doc_id'] for meta in existing_data['metadatas']])
            
        sync_count = 0
        for doc in docs:
            if doc['id'] not in existing_doc_ids:
                # Extract plaintext for vectorization
                content = ""
                if doc['content_encrypted']:
                    decrypted = decrypt(doc['content_encrypted'])
                    if decrypted: content = decrypted
                if not content and doc['content']:
                    content = doc['content']
                if not content and doc['filepath'] and os.path.exists(doc['filepath']):
                    content = extract_text_from_file(doc['filepath'], doc['filename'])
                
                if content:
                    chunks = chunk_text(content)
                    if chunks:
                        ids = [f"doc_{doc['id']}_{i}" for i in range(len(chunks))]
                        metadatas = [{"doc_id": doc['id'], "classification": doc['classification'], "filename": doc['filename']}] * len(chunks)
                        vector_collection.add(documents=chunks, metadatas=metadatas, ids=ids)
                        sync_count += 1
                        
        print(f"Vector Database sync complete. Added {sync_count} new documents to ChromaDB.")
    except Exception as e:
        print(f"Vector DB Sync Error: {e}")

# DATABASE ENCRYPTION (AES-256)
MASTER_PASSWORD = "admin123"

def get_encryption_key():
    # Returns a 32-byte (256-bit) key
    return hashlib.sha256(MASTER_PASSWORD.encode()).digest()

def encrypt(text):
    if not text:
        return ""
    try:
        key = get_encryption_key()
        # Initialize AES in CBC mode (creates a random 16-byte IV)
        cipher = AES.new(key, AES.MODE_CBC)
        iv = cipher.iv
        
        # Pad the text to 16-byte blocks
        padded_data = pad(str(text).encode('utf-8'), AES.block_size)
        ciphertext = cipher.encrypt(padded_data)
        
        # Prepend the IV to the ciphertext and encode to base64
        combined_payload = iv + ciphertext
        return base64.b64encode(combined_payload).decode('utf-8')
    except Exception as e:
        print(f"Encryption error: {e}")
        return str(text)

def decrypt(encrypted_b64):
    if not encrypted_b64:
        return ""
    try:
        key = get_encryption_key()
        # Decode base64
        combined_payload = base64.b64decode(str(encrypted_b64).encode('utf-8'))
        
        # Extract the 16-byte IV and the ciphertext
        iv = combined_payload[:16]
        ciphertext = combined_payload[16:]
        
        # Decrypt and unpad
        cipher = AES.new(key, AES.MODE_CBC, iv)
        decrypted_padded = cipher.decrypt(ciphertext)
        plaintext = unpad(decrypted_padded, AES.block_size).decode('utf-8')
        return plaintext
    except Exception as e:
        # Fallback if the string wasn't valid AES encrypted text
        return str(encrypted_b64)

def require_password(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if session.get('db_password_verified'):
            return f(*args, **kwargs)
        return redirect(url_for('password_page', next=request.url))
    return decorated_function

# DATABASE SETUP 
def get_db():
    conn = sqlite3.connect("secure_ai_chat.db")
    conn.row_factory = sqlite3.Row
    return conn

def execute_with_retry(func, max_retries=5):
    for attempt in range(max_retries):
        try:
            return func()
        except sqlite3.OperationalError as e:
            if "database is locked" in str(e) and attempt < max_retries - 1:
                time.sleep(0.5 * (attempt + 1))
                continue
            raise

def init_db():
    conn = get_db()
    
    conn.execute('''
        CREATE TABLE IF NOT EXISTS company (
            id INTEGER PRIMARY KEY,
            company_name TEXT NOT NULL,
            address TEXT,
            phone TEXT,
            email TEXT,
            registration_number TEXT,
            tax_id TEXT,
            website TEXT,
            established_date TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    conn.execute('''
        CREATE TABLE IF NOT EXISTS departments (
            id INTEGER PRIMARY KEY,
            name TEXT NOT NULL UNIQUE,
            code TEXT,
            description TEXT,
            head_position_id INTEGER,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    conn.execute('''
        CREATE TABLE IF NOT EXISTS positions (
            id INTEGER PRIMARY KEY,
            title TEXT NOT NULL UNIQUE,
            level INTEGER NOT NULL,
            parent_id INTEGER,
            description TEXT,
            access_level TEXT DEFAULT 'Low',
            can_approve_low BOOLEAN DEFAULT 0,
            can_approve_medium BOOLEAN DEFAULT 0,
            can_approve_high BOOLEAN DEFAULT 0,
            can_upload_documents BOOLEAN DEFAULT 0,
            can_manage_users BOOLEAN DEFAULT 0,
            can_view_audit BOOLEAN DEFAULT 0,
            can_view_all_documents BOOLEAN DEFAULT 0,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (parent_id) REFERENCES positions(id)
        )
    ''')
    
    conn.execute('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY,
            username TEXT UNIQUE NOT NULL,
            password TEXT NOT NULL,
            email TEXT NOT NULL,
            position_id INTEGER,
            department_id INTEGER,
            first_name TEXT,
            last_name TEXT,
            ic_number TEXT,
            phone_number TEXT,
            address TEXT,
            bank_name TEXT,
            bank_account_number TEXT,
            bank_account_name TEXT,
            epf_number TEXT,
            socso_number TEXT,
            emergency_contact_name TEXT,
            emergency_contact_phone TEXT,
            hire_date TEXT,
            salary INTEGER,
            is_active BOOLEAN DEFAULT 1,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            updated_at TEXT DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (position_id) REFERENCES positions(id),
            FOREIGN KEY (department_id) REFERENCES departments(id)
        )
    ''')
    
    conn.executescript('''
        CREATE TABLE IF NOT EXISTS documents (
            id INTEGER PRIMARY KEY,
            filename TEXT NOT NULL,
            filepath TEXT NOT NULL,
            classification TEXT NOT NULL CHECK(classification IN ('Low', 'Medium', 'High')),
            content TEXT,
            uploader_id INTEGER,
            uploader_name TEXT,
            upload_date TEXT DEFAULT CURRENT_TIMESTAMP,
            description TEXT,
            file_size INTEGER,
            download_count INTEGER DEFAULT 0,
            recipient_type TEXT DEFAULT 'all',
            restricted_department_id INTEGER,
            FOREIGN KEY (restricted_department_id) REFERENCES departments(id)
        );
        
        CREATE TABLE IF NOT EXISTS document_permissions (
            id INTEGER PRIMARY KEY,
            document_id INTEGER NOT NULL,
            user_id INTEGER NOT NULL,
            can_view BOOLEAN DEFAULT 1,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (document_id) REFERENCES documents(id) ON DELETE CASCADE,
            FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE,
            UNIQUE(document_id, user_id)
        );
        
        CREATE TABLE IF NOT EXISTS audit_logs (
            id INTEGER PRIMARY KEY,
            user_id INTEGER,
            username TEXT,
            action TEXT NOT NULL,
            details TEXT,
            timestamp TEXT DEFAULT CURRENT_TIMESTAMP
        );
        
        CREATE TABLE IF NOT EXISTS access_requests (
            id INTEGER PRIMARY KEY,
            user_id INTEGER NOT NULL,
            username TEXT NOT NULL,
            requester_role TEXT NOT NULL,
            requested_level TEXT NOT NULL CHECK(requested_level IN ('Low', 'Medium', 'High')),
            reason TEXT NOT NULL,
            status TEXT DEFAULT 'pending' CHECK(status IN ('pending', 'approved', 'rejected', 'expired', 'revoked')),
            notes TEXT,
            reviewed_by INTEGER,
            reviewed_by_name TEXT,
            reviewed_at TEXT,
            granted_at TEXT,
            expires_at TEXT,
            timestamp TEXT DEFAULT CURRENT_TIMESTAMP
        );
        
        CREATE TABLE IF NOT EXISTS active_access (
            id INTEGER PRIMARY KEY,
            user_id INTEGER NOT NULL,
            username TEXT NOT NULL,
            granted_level TEXT NOT NULL,
            granted_by INTEGER,
            granted_by_name TEXT,
            granted_at TEXT NOT NULL,
            expires_at TEXT NOT NULL,
            is_active BOOLEAN DEFAULT 1,
            FOREIGN KEY (user_id) REFERENCES users(id)
        );
    ''')
    
    try: conn.execute("ALTER TABLE documents ADD COLUMN recipient_type TEXT DEFAULT 'all'")
    except: pass
    
    try: conn.execute("ALTER TABLE documents ADD COLUMN restricted_department_id INTEGER")
    except: pass
    
    try: conn.execute("ALTER TABLE users ADD COLUMN salary INTEGER")
    except: pass
    
    try:
        conn.execute("ALTER TABLE positions ADD COLUMN can_approve_low BOOLEAN DEFAULT 0")
        conn.execute("ALTER TABLE positions ADD COLUMN can_approve_medium BOOLEAN DEFAULT 0")
        conn.execute("ALTER TABLE positions ADD COLUMN can_approve_high BOOLEAN DEFAULT 0")
        conn.execute("ALTER TABLE positions ADD COLUMN can_view_all_documents BOOLEAN DEFAULT 0")
    except: pass
    
    try: conn.execute("ALTER TABLE documents ADD COLUMN content_encrypted TEXT")
    except: pass
    
    try:
        conn.execute("ALTER TABLE users ADD COLUMN salary_encrypted TEXT")
        conn.execute("ALTER TABLE users ADD COLUMN ic_number_encrypted TEXT")
        conn.execute("ALTER TABLE users ADD COLUMN bank_account_encrypted TEXT")
    except: pass
    
    company_exists = conn.execute("SELECT COUNT(*) FROM company").fetchone()[0]
    
    if company_exists == 0:
        default_positions = [
            ('Admin', 1, None, 'High', 'System Administrator', 1, 1, 1, 1, 1, 1, 1),
            ('CEO', 1, None, 'High', 'Chief Executive Officer', 1, 1, 1, 1, 1, 1, 1),
            ('Manager', 3, None, 'Medium', 'Department Manager', 1, 1, 0, 1, 0, 0, 0),
            ('Employee', 4, None, 'Low', 'Regular Employee', 0, 0, 0, 0, 0, 0, 0),
        ]
        
        for pos in default_positions:
            existing = conn.execute("SELECT id FROM positions WHERE title = ?", (pos[0],)).fetchone()
            if not existing:
                conn.execute('''
                    INSERT INTO positions (title, level, parent_id, access_level, description,
                                          can_approve_low, can_approve_medium, can_approve_high,
                                          can_upload_documents, can_manage_users, can_view_audit, can_view_all_documents)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', pos)
        
        departments = ['Executive', 'IT', 'HR', 'Finance', 'Marketing', 'Operations']
        for dept in departments:
            existing = conn.execute("SELECT id FROM departments WHERE name = ?", (dept,)).fetchone()
            if not existing:
                conn.execute("INSERT INTO departments (name, code, description) VALUES (?, ?, ?)",
                           (dept, dept[:3].upper(), f"{dept} Department"))
        
        conn.commit()
        
        exec_dept = conn.execute("SELECT id FROM departments WHERE name = 'Executive'").fetchone()
        it_dept = conn.execute("SELECT id FROM departments WHERE name = 'IT'").fetchone()
        hr_dept = conn.execute("SELECT id FROM departments WHERE name = 'HR'").fetchone()
        finance_dept = conn.execute("SELECT id FROM departments WHERE name = 'Finance'").fetchone()
        
        admin_pos = conn.execute("SELECT id FROM positions WHERE title = 'Admin'").fetchone()
        ceo_pos = conn.execute("SELECT id FROM positions WHERE title = 'CEO'").fetchone()
        manager_pos = conn.execute("SELECT id FROM positions WHERE title = 'Manager'").fetchone()
        employee_pos = conn.execute("SELECT id FROM positions WHERE title = 'Employee'").fetchone()
        
        #Add dummy IC and Bank data to the end of each user
        test_users = [
            ('admin', 'admin123', 'admin@company.com', admin_pos['id'], exec_dept['id'], 'System', 'Admin', 15000, '800101-14-1234', '1122334455'),
            ('ceo', 'demo123', 'ceo@company.com', ceo_pos['id'], exec_dept['id'], 'John', 'CEO', 50000, '750505-10-5678', '5566778899'),
            ('jason', 'demo123', 'jason@company.com', manager_pos['id'], it_dept['id'], 'Jason', 'Tan', 12000, '851212-01-9012', '2233445566'),
            ('sarah', 'demo123', 'sarah@company.com', manager_pos['id'], hr_dept['id'], 'Sarah', 'Lee', 11500, '880303-05-3456', '9988776655'),
            ('ahmad', 'demo123', 'ahmad@company.com', employee_pos['id'], it_dept['id'], 'Ahmad', 'Bin Ali', 5000, '950707-10-7890', '4455667788'),
            ('linda', 'demo123', 'linda@company.com', employee_pos['id'], finance_dept['id'], 'Linda', 'Wong', 5200, '980909-14-2345', '7788990011'),
        ]
        
        for user in test_users:
            existing = conn.execute("SELECT id FROM users WHERE username = ?", (user[0],)).fetchone()
            if not existing:
                password_hash = generate_password_hash(user[1])
                
                # 2. Insert the raw data (Now includes user[8] for IC and user[9] for Bank)
                conn.execute('''
                    INSERT INTO users (username, password, email, position_id, department_id, 
                                      first_name, last_name, is_active, salary, ic_number, bank_account_number)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (user[0], password_hash, user[2], user[3], user[4], user[5], user[6], 1, user[7], user[8], user[9]))
                
                # 3. Encrypt ALL THREE sensitive fields using AES-256
                encrypted_salary = encrypt(str(user[7]))
                encrypted_ic = encrypt(str(user[8]))
                encrypted_bank = encrypt(str(user[9]))
                
                # 4. Save the AES ciphertexts to the database
                conn.execute("""
                    UPDATE users 
                    SET salary_encrypted = ?, ic_number_encrypted = ?, bank_account_encrypted = ? 
                    WHERE username = ?
                """, (encrypted_salary, encrypted_ic, encrypted_bank, user[0]))
        
        conn.commit()
        print("Default data created successfully")
    
    conn.close()
    print("Database initialized successfully")
    sync_chromadb() # NEW: Sync ChromaDB on startup

# HELPER FUNCTIONS 
def log_activity(user_id, username, action, details=""):
    try:
        def db_operation():
            with get_db() as conn:
                conn.execute("INSERT INTO audit_logs (user_id, username, action, details) VALUES (?, ?, ?, ?)",
                           (user_id, username, action, details))
                conn.commit()
        execute_with_retry(db_operation)
    except Exception as e:
        print(f"Log error: {e}")

def get_user_temporary_access(user_id):
    try:
        def db_operation():
            with get_db() as conn:
                active = conn.execute("""
                    SELECT granted_level FROM active_access 
                    WHERE user_id = ? AND is_active = 1 AND expires_at > datetime('now')
                    ORDER BY 
                        CASE granted_level
                            WHEN 'High' THEN 3
                            WHEN 'Medium' THEN 2
                            WHEN 'Low' THEN 1
                        END DESC
                    LIMIT 1
                """, (user_id,)).fetchone()
                return active['granted_level'] if active else None
        return execute_with_retry(db_operation)
    except Exception:
        return None

def get_user_base_access_level(user_id):
    try:
        def db_operation():
            with get_db() as conn:
                user = conn.execute("""
                    SELECT p.access_level FROM users u
                    JOIN positions p ON u.position_id = p.id
                    WHERE u.id = ?
                """, (user_id,)).fetchone()
                return user['access_level'] if user else 'Low'
        return execute_with_retry(db_operation)
    except Exception:
        return 'Low'

def get_user_department(user_id):
    try:
        def db_operation():
            with get_db() as conn:
                user = conn.execute("""
                    SELECT u.department_id, d.name as department_name
                    FROM users u
                    JOIN departments d ON u.department_id = d.id
                    WHERE u.id = ?
                """, (user_id,)).fetchone()
                return (user['department_id'] if user else None, 
                       user['department_name'] if user else None)
        return execute_with_retry(db_operation)
    except Exception:
        return None, None

def get_effective_access_level(user_id):
    try:
        base_access = get_user_base_access_level(user_id)
        temp_access = get_user_temporary_access(user_id)
        
        level_priority = {'Low': 1, 'Medium': 2, 'High': 3}
        temp_priority = level_priority.get(temp_access, 0)
        base_priority = level_priority.get(base_access, 1)
        
        return temp_access if (temp_access and temp_priority > base_priority) else base_access
    except Exception:
        return 'Low'

def get_user_permissions(user_id):
    try:
        def db_operation():
            with get_db() as conn:
                perms = conn.execute("""
                    SELECT p.can_approve_low, p.can_approve_medium, p.can_approve_high,
                           p.can_upload_documents, p.can_manage_users, p.can_view_audit, p.can_view_all_documents,
                           p.access_level, p.title
                    FROM users u
                    JOIN positions p ON u.position_id = p.id
                    WHERE u.id = ?
                """, (user_id,)).fetchone()
                return dict(perms) if perms else {}
        return execute_with_retry(db_operation)
    except Exception:
        return {}

def get_authorized_docs(access_level, user_id=None):
    if user_id is None:
        user_id = session.get("user_id")
    
    try:
        permissions = get_user_permissions(user_id)
        user_dept_id, _ = get_user_department(user_id)
        
        if permissions.get('can_view_all_documents', False):
            def db_operation_all():
                with get_db() as conn:
                    return conn.execute("SELECT * FROM documents ORDER BY upload_date DESC").fetchall()
            return execute_with_retry(db_operation_all)
        
        level_map = {'Low': ['Low'], 'Medium': ['Low', 'Medium'], 'High': ['Low', 'Medium', 'High']}
        normalized_level = access_level.capitalize() if access_level else 'Low'
        allowed_levels = level_map.get(normalized_level, ['Low'])
        
        def db_operation():
            with get_db() as conn:
                docs = conn.execute("""
                    SELECT DISTINCT d.* FROM documents d
                    WHERE 
                        d.classification IN ({placeholders})
                        AND
                        (
                            (d.recipient_type IS NULL OR d.recipient_type = 'all')
                            OR
                            (d.recipient_type = 'department' AND d.restricted_department_id = ?)
                            OR
                            (EXISTS (SELECT 1 FROM document_permissions dp WHERE dp.document_id = d.id AND dp.user_id = ?))
                            OR
                            d.uploader_id = ?
                        )
                    ORDER BY d.upload_date DESC
                """.format(placeholders=",".join(["?"] * len(allowed_levels))), 
                (*allowed_levels, user_dept_id or -1, user_id, user_id)).fetchall()
                return docs
        
        return execute_with_retry(db_operation)
    except Exception as e:
        print(f"Error getting authorized docs: {e}")
        return []

def cleanup_expired_access():
    """Automatically revoke expired access grants"""
    while True:
        try:
            with get_db() as conn:
                now = datetime.now().isoformat()
                
                result = conn.execute("""
                    UPDATE active_access 
                    SET is_active = 0 
                    WHERE expires_at <= ? AND is_active = 1
                """, (now,))
                revoked = result.rowcount
                
                conn.execute("""
                    UPDATE access_requests 
                    SET status = 'expired' 
                    WHERE status = 'approved' AND expires_at <= ?
                """, (now,))
                
                conn.commit()
                
                if revoked > 0:
                    print(f"[AUTO-REVOKE] {revoked} expired grants at {now}")
                    
        except Exception as e:
            print(f"Cleanup error: {e}")
        
        time.sleep(30)  # Check every 30 seconds

# Start the cleanup thread (add this near the bottom of app.py, before app.run)
cleanup_thread = threading.Thread(target=cleanup_expired_access, daemon=True)
cleanup_thread.start()
print("[SYSTEM] Auto-revoke thread started")

# COMPANY SETUP ROUTES 
@app.route("/setup")
def setup():
    conn = get_db()
    company = conn.execute("SELECT * FROM company").fetchone()
    conn.close()
    if not company:
        return render_template("setup.html")
    return redirect(url_for("login_page"))

@app.route("/api/force_revoke_check", methods=["POST"])
def force_revoke_check():
    """Force check and revoke expired accesses immediately"""
    try:
        if "user_id" not in session:
            return jsonify({"error": "Unauthorized"}), 401
        
        with get_db() as conn:
            now = datetime.now().isoformat()
            
            # Get expired grants before revoking
            expired_grants = conn.execute("""
                SELECT id, user_id, username, granted_level 
                FROM active_access 
                WHERE expires_at <= ? AND is_active = 1
            """, (now,)).fetchall()
            
            # Revoke them
            result = conn.execute("""
                UPDATE active_access 
                SET is_active = 0 
                WHERE expires_at <= ? 
                AND is_active = 1
            """, (now,))
            revoked = result.rowcount
            
            conn.execute("""
                UPDATE access_requests 
                SET status = 'expired' 
                WHERE status = 'approved' 
                AND expires_at <= ?
            """, (now,))
            
            # Log the force check (MOVE THIS BEFORE commit)
            if revoked > 0:
                conn.execute("""
                    INSERT INTO audit_logs (user_id, username, action, details, timestamp)
                    VALUES (?, ?, 'force_revoke_check', ?, ?)
                """, (session["user_id"], session["username"], f"Force revoked {revoked} expired accesses", now))
            
            conn.commit()  # Now commit everything at once
            
            return jsonify({
                "success": True, 
                "revoked": revoked,
                "message": f"Revoked {revoked} expired access grants"
            })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/company/setup", methods=["POST"])
def setup_company():
    try:
        data = request.get_json()
        with get_db() as conn:
            conn.execute("""
                INSERT INTO company (company_name, address, phone, email, registration_number, tax_id, website, established_date)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (data.get('company_name'), data.get('address'), data.get('phone'), data.get('email'),
                  data.get('registration_number'), data.get('tax_id'), data.get('website'), data.get('established_date')))
            conn.commit()
        return jsonify({"success": True, "message": "Company setup completed"})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route("/api/departments", methods=["GET", "POST"])
def manage_departments():
    if request.method == "GET":
        with get_db() as conn:
            departments = conn.execute("SELECT * FROM departments ORDER BY name").fetchall()
            return jsonify([dict(d) for d in departments])
    elif request.method == "POST":
        try:
            data = request.get_json()
            with get_db() as conn:
                conn.execute("INSERT INTO departments (name, code, description) VALUES (?, ?, ?)",
                           (data.get('name'), data.get('code'), data.get('description')))
                conn.commit()
            return jsonify({"success": True})
        except Exception as e:
            return jsonify({"success": False, "error": str(e)}), 500

@app.route("/api/departments/<int:dept_id>", methods=["PUT", "DELETE"])
def update_department(dept_id):
    if request.method == "PUT":
        try:
            data = request.get_json()
            with get_db() as conn:
                conn.execute("UPDATE departments SET name=?, code=?, description=? WHERE id=?",
                           (data.get('name'), data.get('code'), data.get('description'), dept_id))
                conn.commit()
            return jsonify({"success": True})
        except Exception as e:
            return jsonify({"success": False, "error": str(e)}), 500
    elif request.method == "DELETE":
        try:
            with get_db() as conn:
                conn.execute("DELETE FROM departments WHERE id=?", (dept_id,))
                conn.commit()
            return jsonify({"success": True})
        except Exception as e:
            return jsonify({"success": False, "error": str(e)}), 500

@app.route("/api/departments/list", methods=["GET"])
def get_departments_list():
    try:
        if "user_id" not in session:
            return jsonify({"error": "Unauthorized"}), 401
        with get_db() as conn:
            departments = conn.execute("SELECT id, name FROM departments ORDER BY name").fetchall()
        return jsonify([dict(d) for d in departments])
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# POSITIONS API
@app.route("/api/positions", methods=["GET", "POST"])
def manage_positions():
    if request.method == "GET":
        with get_db() as conn:
            positions = conn.execute("""
                SELECT p.*, (SELECT title FROM positions WHERE id = p.parent_id) as parent_title
                FROM positions p ORDER BY p.level, p.title
            """).fetchall()
            return jsonify([dict(p) for p in positions])
    elif request.method == "POST":
        try:
            data = request.get_json()
            with get_db() as conn:
                conn.execute("""
                    INSERT INTO positions (title, level, parent_id, description, access_level,
                                          can_approve_low, can_approve_medium, can_approve_high,
                                          can_upload_documents, can_manage_users, can_view_audit, can_view_all_documents)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    data.get('title'), data.get('level'), data.get('parent_id'), data.get('description'),
                    data.get('access_level', 'Low'),
                    1 if data.get('can_approve_low') else 0,
                    1 if data.get('can_approve_medium') else 0,
                    1 if data.get('can_approve_high') else 0,
                    1 if data.get('can_upload_documents') else 0,
                    1 if data.get('can_manage_users') else 0,
                    1 if data.get('can_view_audit') else 0,
                    1 if data.get('can_view_all_documents') else 0
                ))
                conn.commit()
            return jsonify({"success": True})
        except Exception as e:
            return jsonify({"success": False, "error": str(e)}), 500

@app.route("/api/positions/<int:pos_id>", methods=["PUT", "DELETE"])
def update_position(pos_id):
    if request.method == "PUT":
        try:
            data = request.get_json()
            with get_db() as conn:
                conn.execute("""
                    UPDATE positions 
                    SET title=?, level=?, parent_id=?, description=?, access_level=?,
                        can_approve_low=?, can_approve_medium=?, can_approve_high=?,
                        can_upload_documents=?, can_manage_users=?, can_view_audit=?, can_view_all_documents=?
                    WHERE id=?
                """, (
                    data.get('title'), data.get('level'), data.get('parent_id'), data.get('description'),
                    data.get('access_level', 'Low'),
                    1 if data.get('can_approve_low') else 0,
                    1 if data.get('can_approve_medium') else 0,
                    1 if data.get('can_approve_high') else 0,
                    1 if data.get('can_upload_documents') else 0,
                    1 if data.get('can_manage_users') else 0,
                    1 if data.get('can_view_audit') else 0,
                    1 if data.get('can_view_all_documents') else 0,
                    pos_id
                ))
                conn.commit()
            return jsonify({"success": True})
        except Exception as e:
            return jsonify({"success": False, "error": str(e)}), 500
    elif request.method == "DELETE":
        try:
            with get_db() as conn:
                conn.execute("DELETE FROM positions WHERE id=?", (pos_id,))
                conn.commit()
            return jsonify({"success": True})
        except Exception as e:
            return jsonify({"success": False, "error": str(e)}), 500

# USER REGISTRATION & AUTH 
@app.route("/login")
def login_page():
    if "user_id" in session:
        return redirect(url_for("chat"))
    with get_db() as conn:
        company = conn.execute("SELECT * FROM company").fetchone()
    if not company:
        return redirect(url_for("setup"))
    return render_template("login.html")

@app.route("/register")
def register_page():
    with get_db() as conn:
        company = conn.execute("SELECT * FROM company").fetchone()
    if not company:
        return redirect(url_for("setup"))
    return render_template("register.html")

@app.route("/api/register", methods=["POST"])
def register():
    try:
        data = request.get_json()
        with get_db() as conn:
            existing = conn.execute("SELECT id FROM users WHERE username=?", (data.get('username'),)).fetchone()
            if existing:
                return jsonify({"success": False, "message": "Username already exists"}), 400
            existing_email = conn.execute("SELECT id FROM users WHERE email=?", (data.get('email'),)).fetchone()
            if existing_email:
                return jsonify({"success": False, "message": "Email already registered"}), 400
            
            # --- NEW AES-256 ENCRYPTION LOGIC ---
            raw_salary = data.get('salary', 0)
            raw_ic = data.get('ic_number', '')
            raw_bank = data.get('bank_account_number', '')

            # Encrypt dynamically
            encrypted_salary = encrypt(str(raw_salary)) if raw_salary else None
            encrypted_ic = encrypt(str(raw_ic)) if raw_ic else None
            encrypted_bank = encrypt(str(raw_bank)) if raw_bank else None

            conn.execute("""
                INSERT INTO users (username, password, email, position_id, department_id,
                                  first_name, last_name, ic_number, phone_number, address,
                                  bank_name, bank_account_number, bank_account_name,
                                  epf_number, socso_number, emergency_contact_name,
                                  emergency_contact_phone, hire_date, is_active,
                                  salary, salary_encrypted, ic_number_encrypted, bank_account_encrypted)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1, ?, ?, ?, ?)
            """, (
                data.get('username'), generate_password_hash(data.get('password')), data.get('email'),
                data.get('position_id'), data.get('department_id'),
                data.get('first_name', ''), data.get('last_name', ''), raw_ic,
                data.get('phone_number', ''), data.get('address', ''),
                data.get('bank_name', ''), raw_bank,
                data.get('bank_account_name', ''), data.get('epf_number', ''),
                data.get('socso_number', ''), data.get('emergency_contact_name', ''),
                data.get('emergency_contact_phone', ''), data.get('hire_date', datetime.now().strftime('%Y-%m-%d')),
                raw_salary, encrypted_salary, encrypted_ic, encrypted_bank # <-- Insert the AES ciphertexts!
            ))
            conn.commit()
            new_user = conn.execute("SELECT id, username FROM users WHERE username=?", (data.get('username'),)).fetchone()
        
        log_activity(new_user['id'], new_user['username'], "user_registered", "New user registered")
        return jsonify({"success": True, "message": "Registration successful!"})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500

@app.route("/api/login", methods=["POST"])
def api_login():
    try:
        data = request.get_json()
        username = data.get("username")
        password = data.get("password")
        
        with get_db() as conn:
            user = conn.execute("SELECT * FROM users WHERE username=? AND is_active=1", (username,)).fetchone()
        
        if not user or not check_password_hash(user["password"], password):
            log_activity(None, username, "login_failed", "Invalid credentials")
            return jsonify({"success": False, "message": "Invalid username or password"}), 401
        
        with get_db() as conn:
            position = conn.execute("SELECT title, access_level FROM positions WHERE id=?", (user['position_id'],)).fetchone()
            department = conn.execute("SELECT name FROM departments WHERE id=?", (user['department_id'],)).fetchone()
        
        session.permanent = True
        session["user_id"] = user["id"]
        session["username"] = user["username"]
        session["position"] = position['title'] if position else 'Employee'
        session["department"] = department['name'] if department else 'Unknown'
        session["access_level"] = get_effective_access_level(user["id"])
        
        log_activity(user["id"], user["username"], "login_success", f"Position: {session['position']}")
        return jsonify({"success": True, "redirect": "/chat"})
    except Exception as e:
        return jsonify({"success": False, "message": "Server error"}), 500

@app.route("/logout")
def logout():
    if "user_id" in session:
        log_activity(session["user_id"], session["username"], "logout", "")
    session.clear()
    return redirect(url_for("login_page"))

# PASSWORD PROTECTED VIEWS 
@app.route('/db-password')
def password_page():
    next_url = request.args.get('next', '/view-decrypted')
    return f'''
    <!DOCTYPE html>
    <html>
    <head>
        <title>Database Access - Password Required</title>
        <style>
            body {{
                font-family: monospace;
                background: linear-gradient(135deg, #0f0c29, #302b63);
                min-height: 100vh;
                display: flex;
                justify-content: center;
                align-items: center;
            }}
            .card {{
                background: white;
                padding: 40px;
                border-radius: 20px;
                text-align: center;
                max-width: 400px;
            }}
            input {{
                width: 100%;
                padding: 12px;
                margin: 15px 0;
                border: 2px solid #ddd;
                border-radius: 8px;
                font-size: 16px;
            }}
            button {{
                background: #667eea;
                color: white;
                border: none;
                padding: 12px 30px;
                border-radius: 8px;
                cursor: pointer;
                font-size: 16px;
            }}
            .error {{ color: red; margin-top: 10px; }}
            .hint {{ font-size: 12px; color: #888; margin-top: 20px; }}
        </style>
    </head>
    <body>
        <div class="card">
            <h1>Database Access</h1>
            <p>Enter password to view decrypted data</p>
            <input type="password" id="password" placeholder="Enter master password">
            <button onclick="verify()">Unlock Database</button>
            <div id="error" class="error"></div>
            <div class="hint">
                Without password, you will only see encrypted cipher text
            </div>
        </div>
        <script>
            async function verify() {{
                const password = document.getElementById('password').value;
                const response = await fetch('/verify-db-password', {{
                    method: 'POST',
                    headers: {{'Content-Type': 'application/json'}},
                    body: JSON.stringify({{password: password}})
                }});
                const data = await response.json();
                if (data.success) {{
                    window.location.href = '{next_url}';
                }} else {{
                    document.getElementById('error').textContent = 'Wrong password!';
                }}
            }}
        </script>
    </body>
    </html>
    '''

@app.route('/verify-db-password', methods=['POST'])
def verify_db_password():
    data = request.get_json()
    if data.get('password') == MASTER_PASSWORD:
        session['db_password_verified'] = True
        return jsonify({'success': True})
    return jsonify({'success': False})

@app.route('/view-decrypted')
@require_password
def view_decrypted_data():
    with get_db() as conn:
        docs = conn.execute("SELECT id, filename, content_encrypted FROM documents").fetchall()
        # 1. Update SQL to pull IC and Bank Account
        users = conn.execute("SELECT id, username, salary_encrypted, ic_number_encrypted, bank_account_encrypted FROM users").fetchall()
    
    decrypted_docs = []
    for doc in docs:
        encrypted = doc['content_encrypted']
        decrypted_content = decrypt(encrypted) if encrypted else "No content"
        decrypted_docs.append({
            'id': doc['id'],
            'filename': doc['filename'],
            'content': decrypted_content[:500] if decrypted_content else "No content"
        })
    
    decrypted_users = []
    for user in users:
        # 2. Decrypt all three fields
        encrypted_salary = user['salary_encrypted']
        decrypted_salary = decrypt(encrypted_salary) if encrypted_salary else "None"
        
        encrypted_ic = user['ic_number_encrypted']
        decrypted_ic = decrypt(encrypted_ic) if encrypted_ic else "None"
        
        encrypted_bank = user['bank_account_encrypted']
        decrypted_bank = decrypt(encrypted_bank) if encrypted_bank else "None"

        decrypted_users.append({
            'id': user['id'],
            'username': user['username'],
            'salary': decrypted_salary,
            'ic_number': decrypted_ic,
            'bank_account': decrypted_bank
        })
    
    return f'''
    <!DOCTYPE html>
    <html>
    <head>
        <title>Decrypted Database View</title>
        <style>
            body {{ font-family: monospace; padding: 20px; background: #f0f0f0; }}
            .container {{ max-width: 1400px; margin: 0 auto; }}
            .card {{ background: white; border-radius: 10px; padding: 20px; margin-bottom: 20px; overflow-x: auto; }}
            table {{ width: 100%; border-collapse: collapse; }}
            th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
            th {{ background: #667eea; color: white; }}
            .success {{ background: #d1fae5; padding: 10px; border-radius: 8px; margin-bottom: 20px; }}
        </style>
    </head>
    <body>
        <div class="container">
            <div class="success">
                You have entered the correct password. Here is the DECRYPTED data.
                <a href="/logout-db" style="float: right; color: #dc2626;">Lock Database</a>
            </div>
            
            <div class="card">
                <h2>Documents (Decrypted)</h2>
                <table>
                    <tr><th>ID</th><th>Filename</th><th>Content</th></tr>
                    {''.join([f'<tr><td>{d["id"]}</td><td>{d["filename"]}</td><td>{d["content"][:200]}...</td></tr>' for d in decrypted_docs])}
                </table>
            </div>
            
            <div class="card">
                <h2>Users Sensitive Data (Decrypted)</h2>
                <table>
                    <tr><th>ID</th><th>Username</th><th>Salary (RM)</th><th>IC Number</th><th>Bank Account</th></tr>
                    {''.join([f'<tr><td>{u["id"]}</td><td>{u["username"]}</td><td>{u["salary"]}</td><td>{u["ic_number"]}</td><td>{u["bank_account"]}</td></tr>' for u in decrypted_users])}
                </table>
            </div>
        </div>
    </body>
    </html>
    '''

@app.route('/logout-db')
def logout_db():
    session.pop('db_password_verified', None)
    return redirect(url_for('password_page'))

@app.route('/view-encrypted')
def view_encrypted_data():
    with get_db() as conn:
        docs = conn.execute("SELECT id, filename, content_encrypted FROM documents").fetchall()
        # 1. Update SQL to pull IC and Bank Account
        users = conn.execute("SELECT id, username, salary_encrypted, ic_number_encrypted, bank_account_encrypted FROM users").fetchall()
    
    return f'''
    <!DOCTYPE html>
    <html>
    <head>
        <title>Encrypted Database View</title>
        <style>
            body {{ font-family: monospace; padding: 20px; background: #1a1a2e; color: #0f0; }}
            .container {{ max-width: 1400px; margin: 0 auto; }}
            .card {{ background: #0f0f1a; border-radius: 10px; padding: 20px; margin-bottom: 20px; overflow-x: auto; }}
            table {{ width: 100%; border-collapse: collapse; min-width: 1000px; }}
            th, td {{ border: 1px solid #333; padding: 8px; text-align: left; vertical-align: top; }}
            th {{ background: #333; color: #0f0; }}
            .warning {{ background: #330000; padding: 10px; border-radius: 8px; margin-bottom: 20px; }}
            .encrypted {{ font-family: monospace; font-size: 11px; word-break: break-all; color: #00ff00; }}
        </style>
    </head>
    <body>
        <div class="container">
            <div class="warning">
                WARNING: Without password - You only see ENCRYPTED cipher text
                <a href="/db-password" style="float: right; color: #0f0;">Enter Password to Decrypt</a>
            </div>
            
            <div class="card">
                <h2>Documents (Encrypted)</h2>
                <table>
                    <tr><th>ID</th><th>Filename</th><th>Content (Cipher Text)</th></tr>
                    {''.join([f'<tr><td>{d["id"]}</td><td>{d["filename"]}</td><td class="encrypted">{d["content_encrypted"][:100] if d["content_encrypted"] else "None"}...</td></tr>' for d in docs])}
                </table>
            </div>
            
            <div class="card">
                <h2>Users Sensitive Data (Encrypted)</h2>
                <table>
                    <tr>
                        <th>ID</th>
                        <th>Username</th>
                        <th>Salary (Cipher Text)</th>
                        <th>IC Number (Cipher Text)</th>
                        <th>Bank Account (Cipher Text)</th>
                    </tr>
                    {''.join([f'<tr><td>{u["id"]}</td><td>{u["username"]}</td><td class="encrypted">{u["salary_encrypted"][:45] if u["salary_encrypted"] else "None"}...</td><td class="encrypted">{u["ic_number_encrypted"][:45] if u["ic_number_encrypted"] else "None"}...</td><td class="encrypted">{u["bank_account_encrypted"][:45] if u["bank_account_encrypted"] else "None"}...</td></tr>' for u in users])}
                </table>
            </div>
        </div>
    </body>
    </html>
    '''

# CHAT ROUTE 
@app.route("/chat")
def chat():
    if "user_id" not in session:
        return redirect(url_for("login_page"))
    
    with get_db() as conn:
        company = conn.execute("SELECT * FROM company").fetchone()
    if not company:
        flash("Company setup required.", "error")
        return redirect(url_for("setup"))
    
    current_user_id = session["user_id"]
    current_access_level = get_effective_access_level(current_user_id)
    session["access_level"] = current_access_level
    
    permissions = get_user_permissions(current_user_id)
    _, user_dept_name = get_user_department(current_user_id)
    
    if session.get("username") == "admin":
        return redirect(url_for("admin_management"))
    
    if permissions.get('can_approve_high', False):
        template = "ceo-chat.html"
    elif permissions.get('can_approve_medium', False) or permissions.get('can_approve_low', False):
        template = "manager-chat.html"
    else:
        template = "employee-chat.html"
    
    return render_template(template, 
                          username=session["username"], 
                          position=session.get("position", "Employee"),
                          department=user_dept_name or "Unknown",
                          access_level=current_access_level)

# ACCESS REQUEST SYSTEM 
@app.route("/request_access", methods=["POST"])
def request_access():
    try:
        if "user_id" not in session:
            return jsonify({"success": False, "message": "Unauthorized"}), 401
        
        data = request.get_json()
        requested_level = data.get("requested_level")
        reason = data.get("reason", "").strip()
        
        current_level = get_effective_access_level(session["user_id"])
        level_priority = {'Low': 1, 'Medium': 2, 'High': 3}
        current_priority = level_priority.get(current_level, 1)
        
        if current_priority == 1:
            allowed_request = 'Medium'
        elif current_priority == 2:
            allowed_request = 'High'
        else:
            return jsonify({"success": False, "message": "You already have maximum access"}), 400
        
        if requested_level != allowed_request:
            return jsonify({"success": False, "message": f"You can only request {allowed_request} level access"}), 400
        
        if not reason:
            return jsonify({"success": False, "message": "Reason is required"}), 400
        
        def db_operation():
            with get_db() as conn:
                existing = conn.execute("SELECT * FROM access_requests WHERE user_id=? AND requested_level=? AND status='pending'", 
                                       (session["user_id"], requested_level)).fetchone()
                if existing:
                    return {"error": "You already have a pending request"}
                conn.execute("INSERT INTO access_requests (user_id, username, requester_role, requested_level, reason, status) VALUES (?, ?, ?, ?, ?, 'pending')",
                           (session["user_id"], session["username"], session.get("position", "Employee"), requested_level, reason))
                conn.commit()
                return {"success": True}
        
        result = execute_with_retry(db_operation)
        if "error" in result:
            return jsonify({"success": False, "message": result["error"]}), 400
        
        log_activity(session["user_id"], session["username"], "access_request", f"Requested {requested_level}")
        return jsonify({"success": True, "message": f"{requested_level} access request submitted"})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500

@app.route("/get_access_requests", methods=["GET"])
def get_access_requests():
    try:
        if "user_id" not in session:
            return jsonify({"error": "Unauthorized"}), 401
        
        permissions = get_user_permissions(session["user_id"])
        
        def db_operation():
            with get_db() as conn:
                if permissions.get('can_approve_high', False):
                    requests = conn.execute("SELECT * FROM access_requests WHERE status='pending' ORDER BY timestamp ASC").fetchall()
                elif permissions.get('can_approve_medium', False):
                    requests = conn.execute("""
                        SELECT * FROM access_requests 
                        WHERE user_id = ? OR (status='pending' AND requested_level = 'Medium')
                        ORDER BY timestamp DESC
                    """, (session["user_id"],)).fetchall()
                elif permissions.get('can_approve_low', False):
                    requests = conn.execute("""
                        SELECT * FROM access_requests 
                        WHERE user_id = ? OR (status='pending' AND requested_level = 'Low')
                        ORDER BY timestamp DESC
                    """, (session["user_id"],)).fetchall()
                else:
                    requests = conn.execute("SELECT * FROM access_requests WHERE user_id = ? ORDER BY timestamp DESC", 
                                           (session["user_id"],)).fetchall()
                return requests
        
        requests = execute_with_retry(db_operation)
        return jsonify({"requests": [dict(r) for r in requests], "user_permissions": permissions})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/review_request", methods=["POST"])
def review_request():
    try:
        if "user_id" not in session:
            return jsonify({"error": "Unauthorized"}), 401
        
        permissions = get_user_permissions(session["user_id"])
        data = request.get_json()
        request_id = data.get("request_id")
        decision = data.get("decision")
        notes = data.get("notes", "")
        
        def db_operation():
            with get_db() as conn:
                req = conn.execute("SELECT * FROM access_requests WHERE id=?", (request_id,)).fetchone()
                if not req:
                    return {"error": "Request not found"}
                if req["status"] != "pending":
                    return {"error": f"Request already {req['status']}"}
                
                if req['requested_level'] == 'High' and not permissions.get('can_approve_high', False):
                    return {"error": "Only CEO can approve High level requests"}
                if req['requested_level'] == 'Medium' and not permissions.get('can_approve_medium', False):
                    return {"error": "You don't have permission to approve Medium level requests"}
                if req['requested_level'] == 'Low' and not permissions.get('can_approve_low', False):
                    return {"error": "You don't have permission to approve Low level requests"}
                
                current_time = datetime.now()
                
                if decision == "approved":
                    expires_at = current_time + timedelta(minutes=120)
                    conn.execute("INSERT INTO active_access (user_id, username, granted_level, granted_by, granted_by_name, granted_at, expires_at) VALUES (?, ?, ?, ?, ?, ?, ?)",
                               (req["user_id"], req["username"], req["requested_level"], session["user_id"], session["username"], current_time.isoformat(), expires_at.isoformat()))
                    conn.execute("UPDATE access_requests SET status='approved', notes=?, reviewed_by=?, reviewed_by_name=?, reviewed_at=?, granted_at=?, expires_at=? WHERE id=?",
                               (notes, session["user_id"], session["username"], current_time.isoformat(), current_time.isoformat(), expires_at.isoformat(), request_id))
                    message = f"Access granted to {req['username']} for 2 hours"
                else:
                    conn.execute("UPDATE access_requests SET status='rejected', notes=?, reviewed_by=?, reviewed_by_name=?, reviewed_at=? WHERE id=?",
                               (notes, session["user_id"], session["username"], current_time.isoformat(), request_id))
                    message = f"Request from {req['username']} rejected"
                
                conn.commit()
                return {"success": True, "message": message}
        
        result = execute_with_retry(db_operation)
        if "error" in result:
            return jsonify({"error": result["error"]}), 400
        
        log_activity(session["user_id"], session["username"], "access_granted" if decision == "approved" else "access_rejected", f"{decision} request {request_id}")
        return jsonify({"success": True, "message": result["message"]})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/check_access_status", methods=["GET"])
def check_access_status():
    try:
        if "user_id" not in session:
            return jsonify({"error": "Unauthorized"}), 401
        
        user_id = session["user_id"]
        effective_level = get_effective_access_level(user_id)
        
        def db_operation():
            with get_db() as conn:
                active_grants = conn.execute("SELECT granted_level, granted_at, expires_at FROM active_access WHERE user_id=? AND is_active=1 AND expires_at>datetime('now') ORDER BY expires_at DESC", (user_id,)).fetchall()
                return active_grants
        
        active_grants = execute_with_retry(db_operation)
        grants = []
        for grant in active_grants:
            expires_at = datetime.fromisoformat(grant['expires_at'])
            remaining = expires_at - datetime.now()
            grants.append({'level': grant['granted_level'], 'expires_in_minutes': int(remaining.total_seconds() / 60), 'expires_at': grant['expires_at']})
        
        session["access_level"] = effective_level
        return jsonify({'current_access_level': effective_level, 'base_role': session.get('position', 'Employee'), 'active_grants': grants})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# AUDIT LOGS 
@app.route("/get_audit_logs", methods=["GET"])
def get_audit_logs():
    try:
        if "user_id" not in session:
            return jsonify({"error": "Unauthorized"}), 401
        
        permissions = get_user_permissions(session["user_id"])
        if not permissions.get('can_view_audit', False):
            return jsonify({"error": "Access denied"}), 403
        
        def db_operation():
            with get_db() as conn:
                limit = request.args.get('limit', 100, type=int)
                action_filter = request.args.get('action', None)
                if action_filter:
                    logs = conn.execute("SELECT * FROM audit_logs WHERE action LIKE ? ORDER BY timestamp DESC LIMIT ?", (f'%{action_filter}%', limit)).fetchall()
                else:
                    logs = conn.execute("SELECT * FROM audit_logs ORDER BY timestamp DESC LIMIT ?", (limit,)).fetchall()
                return logs
        
        logs = execute_with_retry(db_operation)
        return jsonify({"logs": [dict(l) for l in logs], "count": len(logs)})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/get_audit_stats", methods=["GET"])
def get_audit_stats():
    try:
        if "user_id" not in session:
            return jsonify({"error": "Unauthorized"}), 401
        
        permissions = get_user_permissions(session["user_id"])
        if not permissions.get('can_view_audit', False):
            return jsonify({"error": "Access denied"}), 403
        
        def db_operation():
            with get_db() as conn:
                stats = {
                    'logins': conn.execute("SELECT COUNT(*) as c FROM audit_logs WHERE action='login_success' AND timestamp>=datetime('now','-7 days')").fetchone()['c'] or 0,
                    'failed_logins': conn.execute("SELECT COUNT(*) as c FROM audit_logs WHERE action='login_failed' AND timestamp>=datetime('now','-7 days')").fetchone()['c'] or 0,
                    'uploads': conn.execute("SELECT COUNT(*) as c FROM audit_logs WHERE action='upload_document' AND timestamp>=datetime('now','-7 days')").fetchone()['c'] or 0,
                    'access_requests': conn.execute("SELECT COUNT(*) as c FROM audit_logs WHERE action='access_request' AND timestamp>=datetime('now','-7 days')").fetchone()['c'] or 0,
                    'access_grants': conn.execute("SELECT COUNT(*) as c FROM audit_logs WHERE action='access_granted' AND timestamp>=datetime('now','-7 days')").fetchone()['c'] or 0,
                    'revokes': conn.execute("SELECT COUNT(*) as c FROM audit_logs WHERE action='revoke_privilege' AND timestamp>=datetime('now','-7 days')").fetchone()['c'] or 0
                }
                return stats
        return jsonify(execute_with_retry(db_operation))
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# PRIVILEGE MANAGEMENT
@app.route("/get_active_privileges", methods=["GET"])
def get_active_privileges():
    try:
        if "user_id" not in session:
            return jsonify({"error": "Unauthorized"}), 401
        
        permissions = get_user_permissions(session["user_id"])
        if not permissions.get('can_manage_users', False):
            return jsonify({"error": "Access denied"}), 403
        
        def db_operation():
            with get_db() as conn:
                grants = conn.execute("""
                    SELECT a.id as grant_id, a.user_id, a.username, a.granted_level, a.granted_by_name, a.expires_at,
                           (SELECT p.title FROM users u JOIN positions p ON u.position_id=p.id WHERE u.id=a.user_id) as position_title
                    FROM active_access a WHERE a.is_active=1 AND a.expires_at>datetime('now') ORDER BY a.expires_at ASC
                """).fetchall()
                return grants
        
        active_grants = execute_with_retry(db_operation)
        privileges = []
        for grant in active_grants:
            expires_at = datetime.fromisoformat(grant['expires_at'])
            remaining = expires_at - datetime.now()
            remaining_minutes = int(remaining.total_seconds() / 60)
            privileges.append({
                'grant_id': grant['grant_id'], 'user_id': grant['user_id'], 'username': grant['username'],
                'position_title': grant['position_title'] or 'Employee', 'granted_level': grant['granted_level'],
                'granted_by': grant['granted_by_name'], 'expires_at': grant['expires_at'],
                'remaining_time': f"{remaining_minutes//60}h {remaining_minutes%60}m" if remaining_minutes >= 60 else f"{remaining_minutes}m"
            })
        return jsonify({"privileges": privileges})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/revoke_privilege", methods=["POST"])
def revoke_privilege():
    try:
        if "user_id" not in session:
            return jsonify({"error": "Unauthorized"}), 401
        
        permissions = get_user_permissions(session["user_id"])
        if not permissions.get('can_manage_users', False):
            return jsonify({"error": "Access denied"}), 403
        
        data = request.get_json()
        grant_id = data.get("grant_id")
        
        def db_operation():
            with get_db() as conn:
                grant = conn.execute("SELECT * FROM active_access WHERE id=?", (grant_id,)).fetchone()
                if not grant:
                    return {"error": "Grant not found"}
                conn.execute("UPDATE active_access SET is_active=0 WHERE id=?", (grant_id,))
                conn.execute("UPDATE access_requests SET status='revoked' WHERE user_id=? AND requested_level=? AND status='approved'", (grant['user_id'], grant['granted_level']))
                conn.commit()
                return {"success": True, "username": grant['username'], "level": grant['granted_level']}
        
        result = execute_with_retry(db_operation)
        if "error" in result:
            return jsonify({"error": result["error"]}), 404
        
        log_activity(session["user_id"], session["username"], "revoke_privilege", f"Revoked {result['level']} from {result['username']}")
        return jsonify({"success": True, "message": f"Privilege revoked for {result['username']}"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# DOCUMENT MANAGEMENT (UPDATED FOR RAG) 
@app.route("/upload_document", methods=["POST"])
def upload_document():
    try:
        if "user_id" not in session:
            return jsonify({"error": "Unauthorized"}), 401
        
        permissions = get_user_permissions(session["user_id"])
        if not permissions.get('can_upload_documents', False):
            return jsonify({"error": "No permission to upload"}), 403
        
        title = request.form.get("title")
        description = request.form.get("description", "")
        classification = request.form.get("classification")
        content = request.form.get("content", "")
        recipient_type = request.form.get("recipient_type", "all")
        restricted_department_id = request.form.get("restricted_department_id")
        selected_users = request.form.getlist("selected_users")
        
        if not title:
            return jsonify({"error": "Title required"}), 400
        
        # Save plaintext separately for Vector Database (ChromaDB)
        plain_content_for_chroma = content
        
        # Encrypt the content before storage in SQLite
        encrypted_content = encrypt(content) if content else ""
        
        # Handle file upload
        uploaded_file = request.files.get("file")
        filepath = ""
        file_size = 0
        
        if uploaded_file and uploaded_file.filename:
            filename = uploaded_file.filename
            filepath = os.path.join(UPLOAD_FOLDER, filename)
            uploaded_file.save(filepath)
            file_size = os.path.getsize(filepath)
            
            extracted_content = extract_text_from_file(filepath, filename)
            if extracted_content:
                encrypted_content = encrypt(extracted_content)
                plain_content_for_chroma = extracted_content
        else:
            if not content:
                content = f"Document: {title}\nDescription: {description}\nClassification: {classification}\nUploaded by: {session['username']}\nDate: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
                encrypted_content = encrypt(content)
                plain_content_for_chroma = content
            
            filename = f"{title.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
            filepath = os.path.join(UPLOAD_FOLDER, filename)
            with open(filepath, 'w', encoding='utf-8') as f:
                f.write(content)
            file_size = len(content)
        
        if file_size > MAX_FILE_SIZE:
            return jsonify({"error": "File too large"}), 400
        
        classification_map = {'low': 'Low', 'medium': 'Medium', 'high': 'High'}
        classification = classification_map.get(classification, 'Low')
        
        def db_operation():
            with get_db() as conn:
                if recipient_type == 'department' and restricted_department_id:
                    cursor = conn.execute("""
                        INSERT INTO documents (filename, filepath, classification, content_encrypted, uploader_id, uploader_name, description, file_size, recipient_type, restricted_department_id)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, (title, filepath, classification, encrypted_content, session["user_id"], session["username"], description, file_size, recipient_type, restricted_department_id))
                else:
                    cursor = conn.execute("""
                        INSERT INTO documents (filename, filepath, classification, content_encrypted, uploader_id, uploader_name, description, file_size, recipient_type)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, (title, filepath, classification, encrypted_content, session["user_id"], session["username"], description, file_size, recipient_type))
                
                doc_id = cursor.lastrowid
                
                if recipient_type == 'specific':
                    for uid in selected_users:
                        if uid and uid != str(session["user_id"]):
                            conn.execute("INSERT INTO document_permissions (document_id, user_id) VALUES (?, ?)", (doc_id, uid))
                    conn.execute("INSERT INTO document_permissions (document_id, user_id) VALUES (?, ?)", (doc_id, session["user_id"]))
                
                conn.commit()
                return doc_id
        
        doc_id = execute_with_retry(db_operation)
        
        # NEW: ADD TO VECTOR DATABASE (CHROMA)
        if plain_content_for_chroma:
            try:
                chunks = chunk_text(plain_content_for_chroma)
                if chunks:
                    ids = [f"doc_{doc_id}_{i}" for i in range(len(chunks))]
                    metadatas = [{"doc_id": doc_id, "classification": classification, "filename": title}] * len(chunks)
                    vector_collection.add(documents=chunks, metadatas=metadatas, ids=ids)
            except Exception as e:
                print(f"Error saving to Vector DB: {e}")
                
        log_activity(session["user_id"], session["username"], "upload_document", f"Uploaded: {title}")
        return jsonify({"success": True, "message": "Document uploaded", "document": {"id": doc_id, "title": title}})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/get_documents", methods=["GET"])
def get_documents():
    try:
        if "user_id" not in session:
            return jsonify({"error": "Unauthorized"}), 401
        
        access_level = get_effective_access_level(session["user_id"])
        docs = get_authorized_docs(access_level, session["user_id"])
        
        documents = []
        for doc in docs:
            with get_db() as conn:
                has_restrictions = conn.execute("SELECT COUNT(*) FROM document_permissions WHERE document_id=?", (doc['id'],)).fetchone()[0] > 0
                is_department_restricted = doc['recipient_type'] == 'department' if doc['recipient_type'] else False
            documents.append({
                'id': doc['id'], 'title': doc['filename'], 'description': doc['description'] or '',
                'classification': doc['classification'].lower(), 'uploader_name': doc['uploader_name'],
                'upload_date': doc['upload_date'], 'download_count': doc['download_count'] or 0,
                'is_restricted': has_restrictions or is_department_restricted
            })
        return jsonify({"documents": documents})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/view_document/<int:doc_id>", methods=["GET"])
def view_document(doc_id):
    try:
        if "user_id" not in session:
            return jsonify({"error": "Unauthorized"}), 401
        
        permissions = get_user_permissions(session["user_id"])
        user_id = session["user_id"]
        user_dept_id, _ = get_user_department(user_id)
        
        def db_operation():
            with get_db() as conn:
                doc = conn.execute("SELECT * FROM documents WHERE id=?", (doc_id,)).fetchone()
                if not doc:
                    return {"error": "Document not found"}
                
                if doc['recipient_type'] == 'department':
                    if doc['restricted_department_id'] != user_dept_id and not permissions.get('can_view_all_documents', False):
                        return {"error": "Access denied. This document is restricted to a specific department."}
                
                has_specific_permissions = conn.execute("SELECT COUNT(*) as count FROM document_permissions WHERE document_id=?", (doc_id,)).fetchone()['count'] > 0
                
                if has_specific_permissions:
                    is_authorized = conn.execute("""
                        SELECT 1 FROM document_permissions WHERE document_id = ? AND user_id = ?
                        UNION ALL
                        SELECT 1 WHERE ? = (SELECT uploader_id FROM documents WHERE id = ?)
                    """, (doc_id, user_id, user_id, doc_id)).fetchone()
                    
                    if not is_authorized and not permissions.get('can_view_all_documents', False):
                        return {"error": "Access denied. This document is restricted to specific users."}
                elif doc['recipient_type'] != 'department':
                    level_map = {'Low': ['Low'], 'Medium': ['Low', 'Medium'], 'High': ['Low', 'Medium', 'High']}
                    access_level = get_effective_access_level(user_id)
                    if doc['classification'] not in level_map.get(access_level, ['Low']):
                        return {"error": f"Access denied. This document requires {doc['classification']} clearance."}
                
                conn.execute("UPDATE documents SET download_count = download_count + 1 WHERE id=?", (doc_id,))
                conn.commit()
                
                # Get content - try encrypted first, then plaintext
                content = ""
                
                try:
                    encrypted_content = doc['content_encrypted']
                    if encrypted_content:
                        decrypted = decrypt(encrypted_content)
                        if decrypted:
                            content = decrypted
                except (KeyError, IndexError):
                    pass
                
                if not content:
                    try:
                        content = doc['content'] or "No content"
                    except (KeyError, IndexError):
                        content = "No content"
                
                if doc['filepath'] and os.path.exists(doc['filepath']):
                    extracted = extract_text_from_file(doc['filepath'], doc['filename'])
                    if extracted:
                        content = extracted
                
                return {
                    "id": doc['id'], "title": doc['filename'], "classification": doc['classification'],
                    "content": content, "uploader": doc['uploader_name'],
                    "upload_date": doc['upload_date'], "description": doc['description'] or '',
                    "views": doc['download_count'] or 0
                }
        
        result = execute_with_retry(db_operation)
        if "error" in result:
            return jsonify({"error": result["error"]}), 403
        log_activity(session["user_id"], session["username"], "view_document", f"Viewed: {result['title']}")
        return jsonify(result)
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

@app.route("/delete_document/<int:doc_id>", methods=["DELETE"])
def delete_document(doc_id):
    try:
        if "user_id" not in session:
            return jsonify({"error": "Unauthorized"}), 401
        
        permissions = get_user_permissions(session["user_id"])
        if not permissions.get('can_manage_users', False):
            return jsonify({"error": "Admin only"}), 403
        
        def db_operation():
            with get_db() as conn:
                doc = conn.execute("SELECT * FROM documents WHERE id=?", (doc_id,)).fetchone()
                if not doc:
                    return {"error": "Not found"}
                if doc['filepath'] and os.path.exists(doc['filepath']):
                    try: os.remove(doc['filepath'])
                    except: pass
                conn.execute("DELETE FROM document_permissions WHERE document_id=?", (doc_id,))
                conn.execute("DELETE FROM documents WHERE id=?", (doc_id,))
                conn.commit()
                return {"success": True, "filename": doc['filename']}
        
        result = execute_with_retry(db_operation)
        if "error" in result:
            return jsonify({"error": result["error"]}), 404
            
        # NEW: DELETE FROM VECTOR DATABASE (CHROMA)
        try:
            vector_collection.delete(where={"doc_id": doc_id})
        except Exception as e:
            print(f"Error deleting from Vector DB: {e}")
            
        log_activity(session["user_id"], session["username"], "delete_document", f"Deleted: {result['filename']}")
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/users/for_selection", methods=["GET"])
def get_users_for_selection():
    try:
        if "user_id" not in session:
            return jsonify({"error": "Unauthorized"}), 401
        
        permissions = get_user_permissions(session["user_id"])
        if not permissions.get('can_upload_documents', False):
            return jsonify({"error": "Unauthorized"}), 403
        
        with get_db() as conn:
            users = conn.execute("""
                SELECT u.id, u.username, u.first_name, u.last_name, p.title as position_title, d.name as department_name
                FROM users u JOIN positions p ON u.position_id=p.id JOIN departments d ON u.department_id=d.id
                WHERE u.is_active=1 AND u.id!=? ORDER BY d.name, p.level, u.username
            """, (session["user_id"],)).fetchall()
        return jsonify([dict(u) for u in users])
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# INTELLIGENT CHAT API (UPDATED FOR RAG)
def extract_text_from_file(filepath, filename):
    ext = filename.lower().split('.')[-1]
    
    try:
        if ext == 'txt':
            with open(filepath, 'r', encoding='utf-8') as f:
                return f.read()
        elif ext == 'pdf':
            text = ""
            with open(filepath, 'rb') as f:
                reader = PyPDF2.PdfReader(f)
                for page in reader.pages:
                    page_text = page.extract_text()
                    if page_text:
                        text += page_text + "\n"
            return text if text else None
        elif ext in ['docx', 'doc']:
            doc = docx.Document(filepath)
            return "\n".join([para.text for para in doc.paragraphs])
        elif ext == 'csv':
            with open(filepath, 'r', encoding='utf-8') as f:
                reader = csv.reader(f)
                return "\n".join([",".join(row) for row in reader])
        else:
            return None
    except Exception as e:
        print(f"Error extracting text: {e}")
        return None

@app.route("/api/chat", methods=["POST"])
def api_chat():
    try:
        if "user_id" not in session:
            return jsonify({"error": "Unauthorized"}), 401
        
        if session.get("username") == "admin":
            return jsonify({"error": "Admin accounts do not have chat access"}), 403
        
        question = request.json.get("question", "").strip()
        referenced_doc_id = request.json.get("doc_id")
        
        if not question:
            return jsonify({"error": "Question cannot be empty"}), 400
        
        username = session["username"]
        position = session.get("position", "Employee")
        access_level = get_effective_access_level(session["user_id"])
        user_id = session["user_id"]
        user_dept_id, user_dept_name = get_user_department(user_id)
        permissions = get_user_permissions(user_id)
        
        log_activity(user_id, username, "chat_query", question[:150])
        
        with get_db() as conn:
            user_info = conn.execute("""
                SELECT u.first_name, u.last_name, u.salary,
                       p.title as pos, d.name as dept
                FROM users u JOIN positions p ON u.position_id=p.id 
                JOIN departments d ON u.department_id=d.id
                WHERE u.id=?
            """, (user_id,)).fetchone()
            
            company = conn.execute("SELECT * FROM company LIMIT 1").fetchone()
            
            dept_salaries = conn.execute("""
                SELECT d.name, AVG(u.salary) as avg, COUNT(u.id) as cnt, 
                       MIN(u.salary) as min, MAX(u.salary) as max, SUM(u.salary) as total
                FROM users u JOIN departments d ON u.department_id=d.id
                WHERE u.salary > 0 GROUP BY d.id
            """).fetchall()
            
            all_active = conn.execute("""
                SELECT u.username, u.first_name, u.last_name, u.salary, 
                       p.title as pos, d.name as dept
                FROM users u JOIN positions p ON u.position_id=p.id 
                JOIN departments d ON u.department_id=d.id
                WHERE u.is_active=1 AND u.salary>0 
                ORDER BY d.name, p.level
            """).fetchall()
            
            own_dept = conn.execute("""
                SELECT u.username, u.first_name, u.last_name, u.salary, p.title as pos
                FROM users u JOIN positions p ON u.position_id=p.id
                WHERE u.department_id=? AND u.is_active=1 AND u.salary>0 
                ORDER BY p.level
            """, (user_dept_id,)).fetchall() if user_dept_id else []
        
        # NEW: VECTOR SEARCH LOGIC
        # Get authorized SQLite Document IDs to ensure RBAC filtering in ChromaDB
        docs = get_authorized_docs(access_level, user_id)
        authorized_doc_ids = [d['id'] for d in docs]
        
        all_docs_content = ""
        has_specific = False
        has_any = False
        doc_context = ""
        can_access_doc = False
        access_denied_reason = ""
        
        # CASE 1: Querying a specific referenced document
        if referenced_doc_id:
            with get_db() as conn:
                doc_data = conn.execute("SELECT * FROM documents WHERE id=?", (referenced_doc_id,)).fetchone()
            
            if doc_data:
                lm = {'Low': 1, 'Medium': 2, 'High': 3}
                doc_classification = doc_data['classification']
                doc_recipient_type = doc_data['recipient_type'] if 'recipient_type' in doc_data.keys() else 'all'
                doc_restricted_dept = doc_data['restricted_department_id'] if 'restricted_department_id' in doc_data.keys() else None
                
                if lm.get(doc_classification, 1) > lm.get(access_level, 1):
                    access_denied_reason = f"requires {doc_classification} clearance (you have {access_level})"
                elif doc_recipient_type == 'department' and doc_restricted_dept != user_dept_id and not permissions.get('can_view_all_documents'):
                    access_denied_reason = "restricted to specific department"
                elif doc_recipient_type == 'specific':
                    with get_db() as conn:
                        auth = conn.execute("""
                            SELECT 1 FROM document_permissions WHERE document_id=? AND user_id=?
                            UNION SELECT 1 WHERE ?=(SELECT uploader_id FROM documents WHERE id=?)
                        """, (referenced_doc_id, user_id, user_id, referenced_doc_id)).fetchone()
                    if not auth and not permissions.get('can_view_all_documents'):
                        access_denied_reason = "restricted to specific users"
                    else: can_access_doc = True
                else: can_access_doc = True
                
                if can_access_doc:
                    # Using Vector DB for specific document search if possible
                    try:
                        results = vector_collection.query(
                            query_texts=[question],
                            n_results=5,
                            where={"doc_id": referenced_doc_id}
                        )
                        if results and results['documents'] and len(results['documents'][0]) > 0:
                            for chunk in results['documents'][0]:
                                doc_context += f"\n{chunk}\n"
                            if doc_context.strip():
                                doc_context = f"\n[PRIMARY DOCUMENT: {doc_data['filename']}]\n{doc_context}\n[END OF DOCUMENT]\n"
                                has_specific = True
                    except Exception as e:
                        print(f"Vector search failed for single doc: {e}")
                        
                    # Fallback to old method if Vector DB didn't return anything
                    if not has_specific:
                        dc = ""
                        try:
                            if doc_data['content_encrypted']:
                                decrypted = decrypt(doc_data['content_encrypted'])
                                if decrypted: dc = decrypted
                        except Exception: pass
                        if not dc and doc_data['content']: dc = doc_data['content']
                        if not dc and doc_data['filepath'] and os.path.exists(doc_data['filepath']):
                            ex = extract_text_from_file(doc_data['filepath'], doc_data['filename'])
                            if ex: dc = ex
                        if dc.strip():
                            doc_context = f"\n[PRIMARY DOCUMENT: {doc_data['filename']}]\n{dc[:4000]}\n[END OF DOCUMENT]\n"
                            has_specific = True
            else:
                doc_context = "[Document not found]"
        
        # CASE 2: General Knowledge Base Search
        elif authorized_doc_ids:
            try:
                # Query ChromaDB specifically for chunks mapped to allowed SQLite doc_ids
                results = vector_collection.query(
                    query_texts=[question],
                    n_results=4, # Bring back top 4 most relevant chunks
                    where={"doc_id": {"$in": authorized_doc_ids}}
                )
                
                if results and results['documents'] and len(results['documents'][0]) > 0:
                    retrieved_chunks = results['documents'][0]
                    retrieved_metadatas = results['metadatas'][0]
                    
                    for chunk, meta in zip(retrieved_chunks, retrieved_metadatas):
                        all_docs_content += f"\n--- Source: {meta.get('filename', 'Unknown')} (Clearance: {meta.get('classification', 'Unknown')}) ---\n{chunk}\n"
                    
                    has_any = bool(all_docs_content)
            except Exception as e:
                print(f"Vector search failed: {e}")
                # Fallback to loading first few documents fully if vector search fails
                if docs:
                    for i, d in enumerate(docs[:3], 1): # Reduced to 3 to prevent overwhelming LLM
                        dc = ""
                        try:
                            if d['content_encrypted']:
                                decrypted = decrypt(d['content_encrypted'])
                                if decrypted: dc = decrypted
                        except Exception: pass
                        if not dc and d['content']: dc = d['content']
                        if not dc and d['filepath'] and os.path.exists(d['filepath']):
                            ex = extract_text_from_file(d['filepath'], d['filename'])
                            if ex: dc = ex
                        if dc.strip():
                            all_docs_content += f"\n--- DOC {i}: {d['filename']} ---\n{dc[:1000]}\n"
                    has_any = bool(all_docs_content)

        
        if access_denied_reason:
            answer = f"I'm sorry, but I cannot access that document. It {access_denied_reason}. Your current access level is {access_level}. If you need access, you can request a temporary elevation from your manager."
            return jsonify({"answer": answer, "access_denied": True, "access_denied_reason": access_denied_reason})
        
        def safe_salary(value):
            try:
                if value is None: return "0"
                return f"{int(float(value)):,}"
            except (ValueError, TypeError):
                return "0"
        
        is_hr = user_dept_name and user_dept_name.lower() == 'hr'
        is_fin = user_dept_name and user_dept_name.lower() == 'finance'
        is_exec = permissions.get('can_view_all_documents', False)
        can_see_all = is_hr or is_fin or is_exec
        
        company_keywords = ['policy', 'company', 'employee', 'hr', 'procedure', 'rule', 'regulation',
                            'dress code', 'working hour', 'leave', 'benefit', 'salary', 'payroll',
                            'our', 'we', 'us', 'department', 'manager', 'office', 'staff', 'team',
                            'attendance', 'holiday', 'vacation', 'sick leave', 'annual leave']
        
        general_keywords = ['what is', 'who is', 'where is', 'when did', 'how does', 'define',
                            'iphone', 'android', 'malaysia', 'indonesia', 'singapore', 'thailand',
                            'france', 'germany', 'japan', 'china', 'america', 'united states',
                            'capital', 'population', 'currency', 'language', 'history', 'science',
                            'math', 'physics', 'chemistry', 'biology', 'computer', 'programming']
        
        greeting_keywords = ['hi', 'hello', 'hey', 'how are you', 'how do you do',
                            'good morning', 'good afternoon', 'good evening',
                            'what\'s up', 'sup', 'howdy', 'greetings',
                            'how you going', 'how are u', 'how u doing', 'how\'s it going']
        
        question_lower = question.lower()
        
        is_greeting = any(keyword in question_lower for keyword in greeting_keywords) or len(question.split()) <= 2
        is_company_question = any(keyword in question_lower for keyword in company_keywords)
        is_general_question = any(keyword in question_lower for keyword in general_keywords)
        
        if can_see_all:
            salary_rule = "FULL ACCESS: You CAN share ALL employee salaries (HR/Finance/Executive privilege)."
            salary_data = "\nALL EMPLOYEE SALARIES (you can share all):\n" + "\n".join([
                f"- {u['first_name'] or ''} {u['last_name'] or ''} (@{u['username']}) - {u['dept']} - {u['pos']} - RM{safe_salary(u['salary'])}" 
                for u in all_active])
        else:
            salary_rule = f"RESTRICTED: Only share {user_dept_name} department salaries. Other departments are FORBIDDEN."
            own_str = "\nYOUR DEPARTMENT SALARIES (can share):\n" + "\n".join([
                f"- {u['first_name'] or ''} {u['last_name'] or ''} (@{u['username']}) - {u['pos']} - RM{safe_salary(u['salary'])}" 
                for u in own_dept]) if own_dept else "\nNo colleagues in your department.\n"
            rest_str = "\nOTHER DEPARTMENT SALARIES (RESTRICTED - DO NOT SHARE):\n" + "\n".join([
                f"- {u['first_name'] or ''} {u['last_name'] or ''} (@{u['username']}) - {u['dept']} - {u['pos']} - [RESTRICTED]" 
                for u in all_active if u['dept'] != user_dept_name])
            salary_data = own_str + rest_str
        
        agg_data = "\nAGGREGATE DATA (available to all users):\n" + "\n".join([
            f"- {d['name']}: {d['cnt']} employees, Avg: RM{safe_salary(d['avg'])}, Range: RM{safe_salary(d['min'])}-RM{safe_salary(d['max'])}, Total: RM{safe_salary(d['total'])}" 
            for d in dept_salaries])
        
        user_salary = safe_salary(user_info['salary']) if user_info else "0"
        
        if is_greeting and not is_company_question:
            prompt = f"""You are a friendly and intelligent AI assistant for {company['company_name'] if company else 'a company'}.

USER CONTEXT:
- Name: {user_info['first_name']} {user_info['last_name']}
- Position: {user_info['pos']}
- Department: {user_info['dept']}
- Access Level: {access_level}

USER MESSAGE: {question}

INSTRUCTIONS:
- This is a greeting or casual conversation starter.
- Be WARM, FRIENDLY, and WELCOMING.
- Address the user by their first name: {user_info['first_name']}.
- Just greet the user and ask how you can help today.
- Keep response to 1-2 sentences maximum.
- Be conversational, not robotic.
- ABSOLUTELY DO NOT ask the user if they want to generate a report, PDF, or Excel file. NEVER ask follow-up questions about files.

YOUR RESPONSE:"""
        elif is_company_question and (has_specific or has_any):
            prompt = f"""You are an expert AI assistant for {company['company_name'] if company else 'the company'}.
You have access to company documents and MUST answer based on the documents provided.

USER CONTEXT:
- Name: {user_info['first_name']} {user_info['last_name']}
- Position: {user_info['pos']}
- Department: {user_info['dept']}
- Access Level: {access_level}

===== COMPANY DOCUMENTS (YOUR ONLY SOURCE FOR COMPANY-SPECIFIC ANSWERS) =====
{doc_context if has_specific else all_docs_content}
===== END OF COMPANY DOCUMENTS =====

USER QUESTION: {question}

STRICT RULES:
1. Answer ONLY using information from the documents above for company-specific content.
2. If the exact policy/procedure is in the document, quote or summarize it accurately.
3. If the document doesn't contain the specific answer, say: "I couldn't find information about [topic] in the available company documents."
4. Be structured and clear in your response.
5. Do NOT make up policies or procedures.
6. ABSOLUTELY DO NOT ask the user if they want to generate a report, PDF, or Excel file. NEVER append questions like "Would you like me to generate a file?". Provide your answer and immediately stop.

YOUR RESPONSE:"""
        elif is_general_question or (not is_company_question and not has_specific and not has_any):
            prompt = f"""You are a helpful and knowledgeable AI assistant. Answer the user's question directly and accurately.

USER: {user_info['first_name']} ({user_info['pos']}, {user_info['dept']})

QUESTION: {question}

INSTRUCTIONS:
- Answer directly using your general knowledge.
- Be accurate, clear, and concise (2-4 sentences).
- If it's a complex topic, give a brief but informative answer.
- If you're not 100% sure, say "I'm not entirely sure, but..." and provide what you know.
- ABSOLUTELY DO NOT ask the user if they want to generate a report, PDF, or Excel file. NEVER ask follow-up questions like "Would you like me to generate a file?". Just answer and stop.

YOUR RESPONSE:"""
        elif has_specific or has_any:
            prompt = f"""You are an AI data analyst for {company['company_name'] if company else 'the company'}.
Analyze the provided data and documents to answer the user's question.

USER CONTEXT:
- Name: {user_info['first_name']} {user_info['last_name']}
- Position: {user_info['pos']}
- Department: {user_info['dept']}
- Access Level: {access_level}
- Salary Access: {salary_rule}
- Your Own Salary: RM {user_salary}

{agg_data}

{salary_data}

{doc_context if has_specific else all_docs_content}

USER QUESTION: {question}

STRICT RULES:
1. Use ONLY the data provided above - do NOT make up numbers.
2. If the data doesn't contain the answer, say: "The available data doesn't contain information about [topic]".
3. For salary-related questions:
   - {salary_rule}
   - Always format salaries as: RM X,XXX
4. If asked to compare or analyze, provide clear comparisons.
5. ABSOLUTELY DO NOT ask the user if they want to generate a report, PDF, or Excel file. NEVER append questions like "Would you like me to generate a file?". Provide your answer and immediately stop.
6. Be precise with numbers and clear in your analysis.

YOUR RESPONSE:"""
        else:
            prompt = f"""You are a helpful AI assistant for {company['company_name'] if company else 'the company'}.

USER: {user_info['first_name']} ({user_info['pos']}, {user_info['dept']})

QUESTION: {question}

INSTRUCTIONS:
- Respond naturally and helpfully.
- If this is about the company but you don't have documents, suggest the user check company resources or ask their manager.
- Keep responses concise and clear (2-4 sentences).
- ABSOLUTELY DO NOT ask the user if they want to generate a report, PDF, or Excel file. NEVER ask follow-up questions like "Would you like me to generate a file?". Just answer and stop.

YOUR RESPONSE:"""
        
        print(f"CHAT | Greeting:{is_greeting} | CompanyQ:{is_company_question} | GeneralQ:{is_general_question} | HasDoc:{has_specific or has_any}")
        
        try:
            response = requests.post("http://localhost:11434/api/generate", json={
                "model": "llama3.2",
                "prompt": prompt,
                "stream": False,
                "temperature": 0.7 if is_greeting else 0.1 if (is_company_question or has_specific or has_any) else 0.7,
                "max_tokens": 500
            }, timeout=90)
            response.raise_for_status()
            answer = response.json().get("response", "Sorry, I couldn't generate a response.")
            answer = answer.replace("**", "").replace("__", "").replace("```", "").strip()
            
            result = {"answer": answer}
            
            wants_report = any(p in question.lower() for p in 
                              ['generate report','generate the report','generate a report', 'create report', 'make report',
                               'generate excel', 'generate pdf', 'create excel', 'create pdf',
                               'download report', 'give me a report', 'excel report', 'pdf report',
                               'yes excel', 'yes pdf', 'yes csv','in pdf', 'in csv'])
            
            if wants_report:
                result["suggest_report"] = True
                if referenced_doc_id:
                    result["doc_id"] = referenced_doc_id
                result["report_context"] = {
                    "ai_answer": answer,
                    "query": question,
                    "user_dept": user_dept_name,
                    "access_level": access_level,
                    "salary_data": salary_data,
                    "agg_data": agg_data
                }
            
            return jsonify(result)
            
        except requests.exceptions.ConnectionError:
            return jsonify({"answer": "AI service is not running. Please start Ollama with 'ollama serve'."})
        except Exception as e:
            return jsonify({"answer": f"AI service temporarily unavailable. Error: {str(e)}"})
            
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

# AI REPORT GENERATION
@app.route("/api/analyze_document_for_report", methods=["POST"])
def analyze_document_for_report():
    try:
        if "user_id" not in session:
            return jsonify({"error": "Unauthorized"}), 401
        
        data = request.get_json()
        doc_id = data.get("doc_id")
        user_query = data.get("query", "")
        
        user_id = session["user_id"]
        access_level = get_effective_access_level(user_id)
        user_dept_id, user_dept_name = get_user_department(user_id)
        permissions = get_user_permissions(user_id)
        
        def db_operation():
            with get_db() as conn:
                return conn.execute("SELECT * FROM documents WHERE id=?", (doc_id,)).fetchone()
        
        doc = execute_with_retry(db_operation)
        if not doc:
            return jsonify({"error": "Document not found"}), 404
        
        if doc['recipient_type'] == 'department' and doc['restricted_department_id'] != user_dept_id:
            if not permissions.get('can_view_all_documents', False):
                return jsonify({"error": "Access denied"}), 403
        
        level_map = {'Low': ['Low'], 'Medium': ['Low', 'Medium'], 'High': ['Low', 'Medium', 'High']}
        if doc['classification'] not in level_map.get(access_level, ['Low']):
            return jsonify({"error": f"Access denied. Requires {doc['classification']} clearance."}), 403
        
        doc_content = ""
        try:
            encrypted_content = doc['content_encrypted']
            if encrypted_content:
                decrypted = decrypt(encrypted_content)
                if decrypted:
                    doc_content = decrypted
        except (KeyError, IndexError):
            pass
        
        if not doc_content:
            try:
                doc_content = doc['content'] or ""
            except (KeyError, IndexError):
                pass
        
        if doc['filepath'] and os.path.exists(doc['filepath']):
            extracted = extract_text_from_file(doc['filepath'], doc['filename'])
            if extracted:
                doc_content = extracted
        
        if not doc_content:
            return jsonify({"error": "Could not extract content"}), 400
        
        with get_db() as conn:
            user_info = conn.execute("""
                SELECT u.first_name, u.last_name, p.title as position_title, d.name as department_name
                FROM users u JOIN positions p ON u.position_id=p.id JOIN departments d ON u.department_id=d.id
                WHERE u.id=?
            """, (user_id,)).fetchone()
        
        prompt = f"""Analyze this document and user query. Suggest if a report should be generated.

USER: {user_info['first_name']} {user_info['last_name']} ({user_info['position_title']}, {user_info['department_name']})
ACCESS LEVEL: {access_level}

DOCUMENT: {doc['filename']}
{doc_content[:4000]}

QUERY: {user_query}

PRIVACY: Only aggregate data can be shared. Individual data restricted to user's own department ({user_info['department_name']}).

Respond in JSON:
{{
    "analysis": "document analysis summary",
    "answer": "answer to query",
    "suggest_report": true/false,
    "report_type": "salary/employee/financial/general",
    "report_title": "title",
    "extracted_data": [{{"label": "Item", "value": "Data"}}]
}}"""

        response = requests.post("http://localhost:11434/api/generate", json={
            "model": "llama3.2", "prompt": prompt, "stream": False,
            "temperature": 0.2, "max_tokens": 800
        }, timeout=120)
        
        result = response.json().get("response", "{}")
        json_match = re.search(r'\{.*\}', result, re.DOTALL)
        if json_match:
            analysis = json.loads(json_match.group())
        else:
            analysis = {
                "analysis": "Document analyzed",
                "answer": result[:500],
                "suggest_report": True,
                "report_type": "general",
                "report_title": f"Analysis of {doc['filename']}",
                "extracted_data": []
            }
        
        log_activity(user_id, session["username"], "analyze_document", f"Analyzed: {doc['filename']}")
        return jsonify({"success": True, "analysis": analysis, "document_id": doc_id})
        
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

def parse_table_from_text(text):
    if not text:
        return None
    
    lines = text.strip().split('\n')
    
    table_lines = []
    for line in lines:
        if '|' in line and not line.strip().startswith('#'):
            table_lines.append(line)
    
    if len(table_lines) < 2:
        return None
    
    headers = [h.strip() for h in table_lines[0].split('|') if h.strip()]
    
    start_idx = 1
    if start_idx < len(table_lines) and any('--' in cell for cell in table_lines[start_idx].split('|')):
        start_idx = 2
    
    rows = []
    for line in table_lines[start_idx:]:
        cells = [cell.strip() for cell in line.split('|') if cell.strip()]
        if len(cells) >= len(headers):
            rows.append(cells[:len(headers)])
    
    if rows:
        return {"headers": headers, "rows": rows}
    
    return None

@app.route("/api/generate_report_from_analysis", methods=["POST"])
def generate_report_from_analysis():
    try:
        if "user_id" not in session:
            return jsonify({"error": "Unauthorized"}), 401
        
        data = request.get_json()
        analysis_data = data.get("analysis_data", {})
        report_type = data.get("report_type", "excel")
        report_title = data.get("title", "Report")
        report_context = data.get("report_context", {})
        
        user_id = session["user_id"]
        username = session["username"]
        user_role = session.get("position", "Employee")
        access_level = get_effective_access_level(user_id)
        user_dept_id, user_dept_name = get_user_department(user_id)
        
        ai_answer = report_context.get("ai_answer", analysis_data.get("answer", ""))
        user_query = report_context.get("query", "")
        
        if not ai_answer:
            return jsonify({"error": "No analysis data available. Please ask a question first."}), 400
        
        query_lower = user_query.lower()
        if 'sales' in query_lower or 'q1' in query_lower or 'revenue' in query_lower:
            dynamic_title = "Sales Report"
        elif 'salary' in query_lower or 'employee' in query_lower or 'staff' in query_lower:
            dynamic_title = "Employee Salary Report"
        elif 'finance' in query_lower or 'budget' in query_lower or 'cost' in query_lower:
            dynamic_title = "Financial Report"
        elif 'department' in query_lower:
            dynamic_title = "Department Summary Report"
        else:
            dynamic_title = report_title
        
        report_data = {
            "title": dynamic_title,
            "generated_by": username,
            "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "sections": []
        }
        
        if user_query:
            report_data["sections"].append({
                "title": "Request",
                "type": "text",
                "content": user_query
            })
        
        report_data["sections"].append({
            "title": "",
            "type": "text",
            "content": ai_answer
        })
        
        log_activity(user_id, username, "generate_report", f"Report: {dynamic_title}")
        
        if report_type == "excel":
            return generate_excel_report(report_data, False)
        else:
            return generate_pdf_report(report_data, False)
            
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

# EXCEL REPORT GENERATION 
def generate_excel_report(report_data, include_charts=True):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    title_font = Font(name='Arial', size=18, bold=True, color='1E293B')
    subtitle_font = Font(name='Arial', size=9, color='64748B')
    section_font = Font(name='Arial', size=12, bold=True, color='3B82F6')
    header_fill = PatternFill(start_color='3B82F6', end_color='3B82F6', fill_type='solid')
    header_font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    cell_font = Font(name='Arial', size=10)
    alt_fill = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )
    
    ws = wb.create_sheet("Summary")
    ws.merge_cells('A1:G1')
    ws['A1'] = report_data.get('title', 'Report')
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:G2')
    ws['A2'] = f"Generated: {report_data.get('generated_at', '')} | By: {report_data.get('generated_by', '')} | Dept: {report_data.get('user_dept', 'N/A')}"
    ws['A2'].font = subtitle_font
    ws['A2'].alignment = Alignment(horizontal='center')
    
    row = 4
    for section in report_data.get('sections', []):
        ws.cell(row=row, column=1, value=f"- {section['title']}").font = Font(name='Arial', size=10)
        row += 1
    
    for section in report_data.get('sections', []):
        sheet_name = section['title'][:31]
        ws = wb.create_sheet(sheet_name)
        
        ws.merge_cells('A1:H1')
        ws['A1'] = section['title']
        ws['A1'].font = section_font
        
        if section.get('type') == 'text':
            ws.merge_cells('A3:H10')
            ws['A3'] = section.get('content', '')
            ws['A3'].font = Font(name='Arial', size=10)
            ws['A3'].alignment = Alignment(wrap_text=True, vertical='top')
        
        elif section.get('headers'):
            row = 3
            for col, header in enumerate(section['headers'], 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border
            
            for data_idx, data_row in enumerate(section.get('rows', [])):
                row += 1
                for col, value in enumerate(data_row, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.font = cell_font
                    cell.border = thin_border
                    if data_idx % 2 == 1:
                        cell.fill = alt_fill
            
            for col in range(1, len(section['headers']) + 1):
                max_len = max((len(str(ws.cell(row=r, column=col).value or '')) for r in range(3, row + 1)), default=10)
                ws.column_dimensions[get_column_letter(col)].width = min(max_len + 2, 35)
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"{report_data.get('title', 'Report').replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

# PDF REPORT GENERATION 
def generate_pdf_report(report_data, include_charts=True):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle('Title', parent=styles['Title'], fontSize=20, textColor=HexColor('#1E293B'))
    section_style = ParagraphStyle('Section', parent=styles['Heading2'], fontSize=13, textColor=HexColor('#3B82F6'))
    
    elements = []
    elements.append(Paragraph(report_data.get('title', 'Report'), title_style))
    elements.append(Paragraph(f"Generated: {report_data.get('generated_at', '')} | By: {report_data.get('generated_by', '')}", styles['Normal']))
    elements.append(Spacer(1, 15))
    
    for section in report_data.get('sections', []):
        elements.append(Paragraph(section['title'], section_style))
        
        if section.get('type') == 'text':
            elements.append(Paragraph(section.get('content', '').replace('\n', '<br/>'), styles['Normal']))
        elif section.get('headers'):
            table_data = [section['headers']] + section['rows']
            col_width = 480 / len(section['headers'])
            t = Table(table_data, colWidths=[col_width] * len(section['headers']))
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), HexColor('#3B82F6')),
                ('TEXTCOLOR', (0, 0), (-1, 0), HexColor('#FFFFFF')),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 8),
                ('FONTSIZE', (0, 1), (-1, -1), 7),
                ('GRID', (0, 0), (-1, -1), 0.5, HexColor('#E2E8F0')),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [HexColor('#FFFFFF'), HexColor('#F8FAFC')]),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('TOPPADDING', (0, 0), (-1, -1), 3),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            ]))
            elements.append(t)
        
        elements.append(Spacer(1, 10))
    
    doc.build(elements)
    buffer.seek(0)
    
    filename = f"{report_data.get('title', 'Report').replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    
    return send_file(buffer, mimetype='application/pdf', as_attachment=True, download_name=filename)

# ADMIN MANAGEMENT 
@app.route("/admin")
def admin_management():
    if "user_id" not in session:
        return redirect(url_for("login_page"))
    
    permissions = get_user_permissions(session["user_id"])
    if not permissions.get('can_manage_users', False):
        flash("Access denied.", "error")
        return redirect(url_for("chat"))
    
    return render_template("admin-management.html", username=session["username"])

@app.route("/api/company", methods=["GET", "PUT"])
def get_company():
    if request.method == "GET":
        try:
            if "user_id" not in session:
                return jsonify({"error": "Unauthorized"}), 401
            with get_db() as conn:
                company = conn.execute("SELECT * FROM company LIMIT 1").fetchone()
            return jsonify(dict(company) if company else {"error": "Not found"})
        except Exception as e:
            return jsonify({"error": str(e)}), 500
    elif request.method == "PUT":
        try:
            if "user_id" not in session:
                return jsonify({"error": "Unauthorized"}), 401
            data = request.get_json()
            with get_db() as conn:
                conn.execute("""
                    UPDATE company SET company_name=?, address=?, phone=?, email=?, 
                    registration_number=?, tax_id=?, website=?, established_date=?
                    WHERE id=(SELECT id FROM company LIMIT 1)
                """, (data.get('company_name'), data.get('address'), data.get('phone'), data.get('email'),
                      data.get('registration_number'), data.get('tax_id'), data.get('website'), data.get('established_date')))
                conn.commit()
            return jsonify({"success": True})
        except Exception as e:
            return jsonify({"error": str(e)}), 500

@app.route("/api/users", methods=["GET", "POST"])
def api_users():
    if request.method == "GET":
        try:
            if "user_id" not in session:
                return jsonify({"error": "Unauthorized"}), 401
            with get_db() as conn:
                users = conn.execute("""
                    SELECT u.id, u.username, u.email, u.first_name, u.last_name, u.phone_number, 
                           u.is_active, u.salary, u.ic_number, u.bank_account_number,
                           p.title as position_title, p.id as position_id, 
                           d.name as department_name, d.id as department_id
                    FROM users u 
                    LEFT JOIN positions p ON u.position_id=p.id 
                    LEFT JOIN departments d ON u.department_id=d.id
                    ORDER BY u.created_at DESC
                """).fetchall()
            return jsonify([dict(u) for u in users])
        except Exception as e:
            return jsonify({"error": str(e)}), 500
            
    elif request.method == "POST":
        try:
            if "user_id" not in session:
                return jsonify({"error": "Unauthorized"}), 401
            data = request.get_json()
            with get_db() as conn:
                existing = conn.execute("SELECT id FROM users WHERE username=?", (data.get('username'),)).fetchone()
                if existing:
                    return jsonify({"success": False, "error": "Username exists"}), 400
                
                password_hash = generate_password_hash(data.get('password', 'password123'))
                
                # Extract the raw data
                raw_salary = data.get('salary', 0)
                raw_ic = data.get('ic_number', '')
                raw_bank = data.get('bank_account_number', '')
                
                # Encrypt dynamically for the database
                encrypted_salary = encrypt(str(raw_salary)) if raw_salary else None
                encrypted_ic = encrypt(str(raw_ic)) if raw_ic else None
                encrypted_bank = encrypt(str(raw_bank)) if raw_bank else None

                conn.execute("""
                    INSERT INTO users (
                        username, password, email, first_name, last_name, phone_number, 
                        department_id, position_id, is_active, 
                        salary, ic_number, bank_account_number,
                        salary_encrypted, ic_number_encrypted, bank_account_encrypted
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    data.get('username'), password_hash, data.get('email'), data.get('first_name', ''),
                    data.get('last_name', ''), data.get('phone_number', ''), data.get('department_id'),
                    data.get('position_id'), 1 if data.get('is_active') else 0, 
                    raw_salary, raw_ic, raw_bank,
                    encrypted_salary, encrypted_ic, encrypted_bank
                ))
                
                conn.commit()
            return jsonify({"success": True})
        except Exception as e:
            return jsonify({"success": False, "error": str(e)}), 500

@app.route("/api/users/<int:user_id>", methods=["PUT", "DELETE"])
def api_user_detail(user_id):
    if request.method == "PUT":
        try:
            if "user_id" not in session:
                return jsonify({"error": "Unauthorized"}), 401
            data = request.get_json()
            
            with get_db() as conn:
                # 1. Fetch the existing user data first so we don't wipe it!
                existing = conn.execute("SELECT * FROM users WHERE id=?", (user_id,)).fetchone()
                
                # 2. Extract salary (Admin might change this)
                raw_salary = data.get('salary', existing['salary'])
                encrypted_salary = encrypt(str(raw_salary)) if raw_salary else None
                
                # 3. Safely get IC and Bank. If Admin didn't send them, keep the old ones!
                raw_ic = data.get('ic_number') if 'ic_number' in data else existing['ic_number']
                encrypted_ic = encrypt(str(raw_ic)) if raw_ic else None
                
                raw_bank = data.get('bank_account_number') if 'bank_account_number' in data else existing['bank_account_number']
                encrypted_bank = encrypt(str(raw_bank)) if raw_bank else None
                
                if data.get('password'):
                    password_hash = generate_password_hash(data['password'])
                    conn.execute("""
                        UPDATE users SET 
                            username=?, password=?, email=?, first_name=?, last_name=?, 
                            phone_number=?, department_id=?, position_id=?, is_active=?, 
                            salary=?, salary_encrypted=?,
                            ic_number=?, ic_number_encrypted=?,
                            bank_account_number=?, bank_account_encrypted=?
                        WHERE id=?
                    """, (
                        data.get('username'), password_hash, data.get('email'), data.get('first_name', ''),
                        data.get('last_name', ''), data.get('phone_number', ''), data.get('department_id'),
                        data.get('position_id'), 1 if data.get('is_active') else 0, 
                        raw_salary, encrypted_salary,
                        raw_ic, encrypted_ic,
                        raw_bank, encrypted_bank,
                        user_id
                    ))
                else:
                    conn.execute("""
                        UPDATE users SET 
                            username=?, email=?, first_name=?, last_name=?, phone_number=?,
                            department_id=?, position_id=?, is_active=?, 
                            salary=?, salary_encrypted=?,
                            ic_number=?, ic_number_encrypted=?,
                            bank_account_number=?, bank_account_encrypted=?
                        WHERE id=?
                    """, (
                        data.get('username'), data.get('email'), data.get('first_name', ''),
                        data.get('last_name', ''), data.get('phone_number', ''), data.get('department_id'),
                        data.get('position_id'), 1 if data.get('is_active') else 0, 
                        raw_salary, encrypted_salary,
                        raw_ic, encrypted_ic,
                        raw_bank, encrypted_bank,
                        user_id
                    ))
                
                conn.commit()
            return jsonify({"success": True})
        except Exception as e:
            return jsonify({"success": False, "error": str(e)}), 500
            
    elif request.method == "DELETE":
        try:
            if "user_id" not in session:
                return jsonify({"error": "Unauthorized"}), 401
            if user_id == session["user_id"]:
                return jsonify({"success": False, "error": "Cannot delete yourself"}), 400
            with get_db() as conn:
                conn.execute("DELETE FROM users WHERE id=?", (user_id,))
                conn.commit()
            return jsonify({"success": True})
        except Exception as e:
            return jsonify({"success": False, "error": str(e)}), 500
            
    elif request.method == "DELETE":
        try:
            if "user_id" not in session:
                return jsonify({"error": "Unauthorized"}), 401
            if user_id == session["user_id"]:
                return jsonify({"success": False, "error": "Cannot delete yourself"}), 400
            with get_db() as conn:
                conn.execute("DELETE FROM users WHERE id=?", (user_id,))
                conn.commit()
            return jsonify({"success": True})
        except Exception as e:
            return jsonify({"success": False, "error": str(e)}), 500

# START THE APP 
if __name__ == "__main__":
    print("\n" + "="*75)
    print("SECURE AI CHAT SYSTEM - WITH DATABASE ENCRYPTION")
    print("="*75)
    
    init_db()
    
    print("\nDefault Users:")
    print("  admin / admin123 - System Admin")
    print("  ceo / ceo123 - CEO")
    print("  jason / demo123 - IT Manager")
    print("  sarah / demo123 - HR Manager")
    print("  ahmad / demo123 - IT Employee")
    print("  linda / demo123 - Finance Employee")
    
    print("\nDatabase Encryption:")
    print("  Master Password: admin123")
    print("  View encrypted data: /view-encrypted")
    print("  View decrypted data: /view-decrypted (password required)")
    
    print("\nMake sure Ollama is running: ollama serve")
    print("="*75 + "\n")
    
    app.run(host="0.0.0.0", port=5000, debug=True)